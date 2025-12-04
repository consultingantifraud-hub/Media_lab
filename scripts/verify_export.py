#!/usr/bin/env python3
"""Verify export has all required columns."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.export_statistics_to_excel import export_statistics_to_excel
from openpyxl import load_workbook
import tempfile
import os

# Create test export
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    test_file = tmp.name

try:
    print("Creating export...")
    export_statistics_to_excel(test_file)
    
    print(f"\nReading file: {test_file}")
    wb = load_workbook(test_file)
    
    # Check "Статистика по дням"
    if "Статистика по дням" in wb.sheetnames:
        ws = wb["Статистика по дням"]
        headers = [c.value for c in ws[1]]
        print(f"\n📊 Статистика по дням:")
        print(f"   Колонки: {headers}")
        expected = ["Дата", "Количество операций", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)", "Уникальных пользователей"]
        if headers == expected:
            print("   ✅ Все колонки на месте!")
        else:
            print(f"   ❌ Ожидалось: {expected}")
    
    # Check "Статистика по пользователям"
    if "Статистика по пользователям" in wb.sheetnames:
        ws = wb["Статистика по пользователям"]
        headers = [c.value for c in ws[1]]
        print(f"\n📊 Статистика по пользователям:")
        print(f"   Колонки: {headers}")
        expected = ["Telegram ID", "Username", "Имя", "Тип операции", "Модель", "Количество", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)"]
        if headers == expected:
            print("   ✅ Все колонки на месте!")
        else:
            print(f"   ❌ Ожидалось: {expected}")
    
    # Check "Сводка"
    if "Сводка" in wb.sheetnames:
        ws = wb["Сводка"]
        print(f"\n📊 Сводка:")
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if row[0]:
                rows.append(f"   {row[0]}: {row[1]}")
        print("\n".join(rows))
        if any("Общая себестоимость" in r for r in rows) and any("Общая прибыль" in r for r in rows):
            print("   ✅ Себестоимость и прибыль присутствуют!")
        else:
            print("   ❌ Себестоимость или прибыль отсутствуют!")
    
finally:
    if os.path.exists(test_file):
        os.remove(test_file)





