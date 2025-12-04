#!/usr/bin/env python3
"""Export user statistics to Excel file."""
import sys
from pathlib import Path
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.db.base import SessionLocal
from app.db.models import User, UserStatistics, Operation, Balance, AiAssistantQuestion
from app.services.pricing import get_operation_name
from app.core.config import settings
from sqlalchemy import func, desc, extract, case
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta, timezone
from collections import defaultdict

# Moscow timezone (UTC+3)
MOSCOW_TZ = timezone(timedelta(hours=3))

def to_moscow_time(dt: datetime | None) -> datetime | None:
    """Convert datetime to Moscow timezone (UTC+3)."""
    if dt is None:
        return None
    # If datetime is naive (no timezone), assume it's UTC
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    # Convert to Moscow time
    return dt.astimezone(MOSCOW_TZ)

def format_datetime_moscow(dt: datetime | None, format_str: str = "%d.%m.%Y %H:%M") -> str:
    """Format datetime in Moscow timezone."""
    if dt is None:
        return ""
    moscow_dt = to_moscow_time(dt)
    return moscow_dt.strftime(format_str)

# Date when we started storing prices in kopecks
# All operations before this date are assumed to be in rubles (old format)
# All operations after this date are assumed to be in kopecks (new format)
# Based on DB analysis: migration happened around 2025-11-25 09:30 UTC
KOPECKS_MIGRATION_DATETIME = datetime(2025, 11, 25, 9, 30, 0, tzinfo=timezone.utc)

# Exchange rate: 1 USD = 90 RUB
USD_TO_RUB_RATE = 90.0

# Model costs in USD (from fal.ai pricing)
MODEL_COSTS_USD = {
    "fal-ai/nano-banana-pro": 0.15,
    "fal-ai/nano-banana-pro/edit": 0.15,
    "fal-ai/nano-banana": 0.0398,
    "fal-ai/nano-banana/edit": 0.0398,
    "fal-ai/flux-2-flex": 0.06,  # Flux 2 Flex - 6 центов за операцию
    "fal-ai/flux-2-pro/edit": 0.10,  # Flux 2 Pro Edit - примерная стоимость (временно отключена, но может быть в старых данных)
    "fal-ai/bytedance/seedream/v4/edit": 0.03,  # Обратная совместимость
    "fal-ai/bytedance/seedream/v4/text-to-image": 0.03,  # Обратная совместимость
    settings.fal_seedream_edit_model: 0.03,  # Текущая версия (v4.5)
    settings.fal_seedream_create_model: 0.03,  # Текущая версия (v4.5)
    "fal-ai/any-llm": 0.001,
    "fal-ai/recraft/upscale/crisp": 0.004,
    "fal-ai/retoucher": 0.0013,
    "fal-ai/face-swap": 0.01,  # Базовая модель Face Swap от Fal.ai
    "fal-ai/chrono-edit": 0.05,  # Chrono Edit - примерная стоимость
    "wavespeed-ai/image-face-swap": 0.01,
    "openai/gpt-image-1-mini": 0.15,  # OpenAI GPT Image 1 Mini через WaveSpeedAI (аналогично nano-banana-pro)
    "openai/gpt-image-1-mini/edit": 0.15,  # OpenAI GPT Image 1 Mini Edit через WaveSpeedAI (аналогично nano-banana-pro/edit)
}

def get_model_by_operation_type(operation_type: str) -> str | None:
    """Get default model name for operation type if model is not specified."""
    # Маппинг типов операций на модели по умолчанию
    operation_to_model = {
        "upscale": "fal-ai/recraft/upscale/crisp",  # Улучшение качества - $0.004
        "retouch": settings.fal_seedream_edit_model,  # Ретушь - используем Seedream модель ($0.03)
        "face_swap": "fal-ai/face-swap",  # Замена лица - $0.01 (базовая модель, может быть wavespeed-ai/image-face-swap)
        "add_text": "openai/gpt-image-1-mini/edit",  # Добавление текста (через WaveSpeedAI)
        "prompt_generation": "fal-ai/any-llm",  # Генерация промпта - $0.001
    }
    return operation_to_model.get(operation_type)


def get_model_cost_rub(model: str | None, operation_type: str | None = None) -> float:
    """Get model cost in rubles. Returns 0 if model is None or not found.
    
    If model is None or empty, tries to determine model from operation_type.
    """
    # Если модель не указана, пробуем определить по типу операции
    if not model and operation_type:
        model = get_model_by_operation_type(operation_type)
    
    if not model:
        return 0.0
    # Нормализуем имя модели (убираем пробелы, приводим к нижнему регистру для сравнения)
    model_normalized = model.strip().lower() if isinstance(model, str) else str(model).lower()
    # Пробуем найти точное совпадение (приводим ключи словаря к нижнему регистру для сравнения)
    cost_usd = 0.0
    matched_key = None
    for key, value in MODEL_COSTS_USD.items():
        if key.lower() == model_normalized:
            cost_usd = value
            matched_key = key
            break
    # Если не найдено точное совпадение, пробуем найти по частичному совпадению
    if cost_usd == 0.0:
        for key, value in MODEL_COSTS_USD.items():
            key_lower = key.lower()
            # Проверяем, содержит ли нормализованная модель ключ или наоборот
            if model_normalized in key_lower or key_lower in model_normalized:
                cost_usd = value
                matched_key = key
                break
    # Логируем для отладки (только для важных моделей или если не найдено)
    if cost_usd == 0.0 and model_normalized and ('flux' in model_normalized or 'nano' in model_normalized or 'seedream' in model_normalized):
        import sys
        # Логируем только если это важная модель, но не найдена
        print(f"WARNING: get_model_cost_rub: Model '{model}' (normalized: '{model_normalized}') not found in MODEL_COSTS_USD", file=sys.stderr, flush=True)
    result = cost_usd * USD_TO_RUB_RATE
    # DEBUG: Log for flux-2-flex
    if model and 'flux-2-flex' in model.lower():
        import sys
        print(f"DEBUG get_model_cost_rub: model={repr(model)}, normalized={model_normalized}, cost_usd={cost_usd}, result={result}", file=sys.stderr, flush=True)
    return result


def convert_price_to_rubles(price: int | None, created_at: datetime | None) -> float:
    """
    Convert price from kopecks to rubles.
    
    ASSUMPTION: All operations before KOPECKS_MIGRATION_DATETIME are in rubles (old format).
                All operations after KOPECKS_MIGRATION_DATETIME are in kopecks (new format).
    
    If created_at is None, uses heuristic: price > 100 is likely kopecks.
    """
    if price is None or price == 0:
        return 0.0
    
    # If we have created_at, use it to determine format based on migration date
    if created_at is not None:
        # SQLite returns naive datetime (no timezone), assume UTC
        if created_at.tzinfo is None:
            created_at_utc = created_at.replace(tzinfo=timezone.utc)
        else:
            created_at_utc = created_at.astimezone(timezone.utc)
        
        # Simple rule: before migration date = rubles, after = kopecks
        if created_at_utc < KOPECKS_MIGRATION_DATETIME:
            # Old format: price is already in rubles, return as is
            return float(price)
        else:
            # New format: price is in kopecks, convert to rubles
            return float(price) / 100.0
    
    # Fallback heuristic if created_at is None
    # Prices > 100 are likely in kopecks (e.g., 800 for 8₽, 2600 for 26₽)
    # Prices <= 100 are likely in rubles (e.g., 8 for 8₽, 24 for 24₽)
    if price > 100:
        return float(price) / 100.0
    else:
        return float(price)


def export_statistics_to_excel(output_file: str = "statistics_export.xlsx"):
    """Export all statistics to Excel file."""
    db = SessionLocal()
    try:
        output_path = Path(output_file)
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Users sheet
        ws_users = wb.create_sheet("Пользователи")
        headers = ["ID", "Telegram ID", "Username", "Имя", "Фамилия", "Язык", "Premium", 
                  "Регистрация", "Последняя активность", "Баланс", "Всего операций", 
                  "Всего потрачено", "Первая операция", "Последняя операция"]
        ws_users.append(headers)
        
        # Style headers
        for cell in ws_users[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        users = db.query(User).order_by(desc(User.created_at)).all()
        
        for user in users:
            stats = db.query(UserStatistics).filter(UserStatistics.user_id == user.id).first()
            balance = db.query(Balance).filter(Balance.user_id == user.id).first()
            
            # Recalculate total_spent only for operations after migration date (to sync with other sheets)
            total_spent_after_migration = 0.0
            total_operations_after_migration = 0
            first_operation_after_migration = None
            last_operation_after_migration = None
            
            if stats:
                # Get all charged/free operations for this user after migration date
                user_ops = db.query(Operation).filter(
                    Operation.user_id == user.id,
                    Operation.status.in_(["charged", "free"])
                ).all()
                
                for op in user_ops:
                    if op.created_at:
                        if op.created_at.tzinfo is None:
                            created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                        else:
                            created_at_utc = op.created_at.astimezone(timezone.utc)
                        
                        if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                            total_operations_after_migration += 1
                            if op.status == "charged":
                                total_spent_after_migration += float(op.price) / 100.0
                            
                            if first_operation_after_migration is None or op.created_at < first_operation_after_migration:
                                first_operation_after_migration = op.created_at
                            if last_operation_after_migration is None or op.created_at > last_operation_after_migration:
                                last_operation_after_migration = op.created_at
                    elif op.price > 100:  # Heuristic: likely kopecks
                        total_operations_after_migration += 1
                        if op.status == "charged":
                            total_spent_after_migration += float(op.price) / 100.0
            
            ws_users.append([
                user.id,
                user.telegram_id,
                f"@{user.username}" if user.username else "",
                user.first_name or "",
                user.last_name or "",
                user.language_code or "",
                "Да" if user.is_premium else "Нет",
                format_datetime_moscow(user.created_at),
                format_datetime_moscow(user.last_activity_at),
                (balance.balance / 100.0) if balance else 0.0,  # Баланс хранится в копейках, конвертируем в рубли
                total_operations_after_migration,
                total_spent_after_migration,
                format_datetime_moscow(first_operation_after_migration),
                format_datetime_moscow(last_operation_after_migration),
            ])
        
        # 2. Operations by type sheet
        ws_ops_type = wb.create_sheet("Операции по типам")
        ws_ops_type.append(["Тип операции", "Количество", "Выручка (₽)"])
        for cell in ws_ops_type[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all operations and aggregate in Python
        # Only count revenue from operations after kopecks migration (ignore old data)
        all_ops = db.query(Operation).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        
        ops_by_type = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
        for op in all_ops:
            # Only count operations after migration date
            is_after_migration = False
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif op.price > 100:
                is_after_migration = True
            
            if is_after_migration:
                ops_by_type[op.type]['count'] += 1
                if op.status == "charged":
                    price_rubles = float(op.price) / 100.0
                    ops_by_type[op.type]['revenue'] += price_rubles
        
        # Sort by count descending
        sorted_ops = sorted(ops_by_type.items(), key=lambda x: x[1]['count'], reverse=True)
        
        for op_type, data in sorted_ops:
            ws_ops_type.append([get_operation_name(op_type), data['count'], data['revenue']])
        
        # 3. Models used sheet
        ws_models = wb.create_sheet("Использованные модели")
        ws_models.append(["Модель", "Количество использований", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)"])
        for cell in ws_models[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all operations with models and aggregate in Python
        # Only count revenue from operations after kopecks migration
        all_ops_with_models = db.query(Operation).filter(
            Operation.status.in_(["charged", "free"]),
            Operation.model.isnot(None)
        ).all()
        
        models_data = defaultdict(lambda: {'count': 0, 'revenue': 0.0, 'cost': 0.0})
        for op in all_ops_with_models:
            # Only count operations after migration date
            is_after_migration = False
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif op.price > 100:
                is_after_migration = True
            
            if is_after_migration:
                models_data[op.model]['count'] += 1
                # Calculate cost for all operations (charged and free) - cost is real expense
                # Если модель не указана, определяем по типу операции
                model_cost = get_model_cost_rub(op.model, op.type)
                # DEBUG: Log cost calculation for flux-2-flex
                if op.model and 'flux-2-flex' in op.model.lower():
                    import sys
                    print(f"DEBUG models: model={repr(op.model)}, cost={model_cost}, count={models_data[op.model]['count']}, before_add={models_data[op.model]['cost']}", file=sys.stderr, flush=True)
                models_data[op.model]['cost'] += model_cost
                # DEBUG: Log after adding
                if op.model and 'flux-2-flex' in op.model.lower():
                    import sys
                    print(f"DEBUG models AFTER: model={repr(op.model)}, cost_after={models_data[op.model]['cost']}", file=sys.stderr, flush=True)
                if op.status == "charged":
                    price_rubles = float(op.price) / 100.0
                    models_data[op.model]['revenue'] += price_rubles
        
        # Sort by count descending
        sorted_models = sorted(models_data.items(), key=lambda x: x[1]['count'], reverse=True)
        
        for model, data in sorted_models:
            profit = data['revenue'] - data['cost']
            ws_models.append([model, data['count'], data['revenue'], data['cost'], profit])
        
        # 4. All operations sheet
        ws_operations = wb.create_sheet("Все операции")
        ws_operations.append(["ID операции", "Telegram ID", "Тип", "Модель", "Цена", 
                             "Статус", "Дата", "Промпт", "Количество изображений"])
        for cell in ws_operations[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Only show operations after migration date
        all_ops_for_list = db.query(Operation).filter(
            Operation.status.in_(["charged", "free"])
        ).order_by(desc(Operation.created_at)).all()
        
        operations = []
        for op in all_ops_for_list:
            # Only include operations after migration date
            is_after_migration = False
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif op.price > 100:
                is_after_migration = True
            
            if is_after_migration:
                operations.append(op)
        
        for op in operations:
            user = db.query(User).filter(User.id == op.user_id).first()
            prompt = (op.prompt[:200] + "...") if op.prompt and len(op.prompt) > 200 else (op.prompt or "")
            # Convert price - only for operations after migration date
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    price_rubles = float(op.price) / 100.0
                else:
                    price_rubles = float(op.price)  # Old format, but show as is
            elif op.price > 100:
                price_rubles = float(op.price) / 100.0
            else:
                price_rubles = float(op.price)
            ws_operations.append([
                op.id,
                user.telegram_id if user else "",
                get_operation_name(op.type),
                op.model or "",
                price_rubles,
                op.status,
                format_datetime_moscow(op.created_at),
                prompt,
                op.image_count or ""
            ])
        
        # 5. Summary sheet
        ws_summary = wb.create_sheet("Сводка")
        ws_summary.append(["Параметр", "Значение"])
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        total_users = db.query(func.count(User.id)).scalar()
        # Count only operations after migration date
        all_ops_count = db.query(Operation).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        total_operations = 0
        for op in all_ops_count:
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    total_operations += 1
            elif op.price > 100:
                total_operations += 1
        # Calculate total revenue and cost - only from operations after kopecks migration
        all_charged_ops = db.query(Operation).filter(
            Operation.status == "charged"
        ).all()
        total_revenue = 0.0
        total_cost = 0.0
        for op in all_charged_ops:
            # Only count revenue from operations after migration date
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    # Calculate cost for all operations (charged and free) - cost is real expense
                    # Если модель не указана, определяем по типу операции
                    total_cost += get_model_cost_rub(op.model, op.type)
                    if op.status == "charged":
                        total_revenue += float(op.price) / 100.0
            elif op.price > 100:
                # Calculate cost for all operations (charged and free) - cost is real expense
                # Если модель не указана, определяем по типу операции
                total_cost += get_model_cost_rub(op.model, op.type)
                if op.status == "charged":
                    total_revenue += float(op.price) / 100.0
        total_balance_kopecks = db.query(func.sum(Balance.balance)).scalar() or 0
        total_balance_rubles = float(total_balance_kopecks) / 100.0  # Баланс хранится в копейках, конвертируем в рубли
        total_profit = total_revenue - total_cost
        
        ws_summary.append(["Всего пользователей", total_users])
        ws_summary.append(["Всего операций", total_operations])
        ws_summary.append(["Всего заработано (₽)", total_revenue])
        ws_summary.append(["Общая себестоимость (₽)", total_cost])
        ws_summary.append(["Общая прибыль (₽)", total_profit])
        ws_summary.append(["Общий баланс пользователей (₽)", total_balance_rubles])
        ws_summary.append(["Дата выгрузки", format_datetime_moscow(datetime.now(timezone.utc))])
        
        # 6. User operations statistics sheet
        ws_user_ops = wb.create_sheet("Статистика по пользователям")
        ws_user_ops.append(["Telegram ID", "Username", "Имя", "Тип операции", "Модель", "Количество", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)"])
        for cell in ws_user_ops[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all operations with users and models, aggregate in Python to handle date-based conversion
        all_user_ops = db.query(
            User.telegram_id,
            User.username,
            User.first_name,
            Operation.type,
            Operation.model,
            Operation.price,
            Operation.status,
            Operation.created_at
        ).join(
            Operation, User.id == Operation.user_id
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        
        user_ops_data = defaultdict(lambda: {'count': 0, 'revenue': 0.0, 'cost': 0.0})
        for tg_id, username, first_name, op_type, model, price, status, created_at in all_user_ops:
            # Only count operations after migration date
            is_after_migration = False
            if created_at:
                if created_at.tzinfo is None:
                    created_at_utc = created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif price > 100:
                is_after_migration = True
            
            if is_after_migration:
                # Key includes model to track which models were used
                key = (tg_id, username, first_name, op_type, model or "")
                user_ops_data[key]['count'] += 1
                # Calculate cost for all operations (charged and free) - cost is real expense
                # Если модель не указана, определяем по типу операции
                model_cost = get_model_cost_rub(model, op_type)
                # DEBUG: Log cost calculation for flux-2-flex
                if model and 'flux-2-flex' in model.lower():
                    import sys
                    print(f"DEBUG user_ops: model={repr(model)}, cost={model_cost}, count={user_ops_data[key]['count']}, before_add={user_ops_data[key]['cost']}", file=sys.stderr, flush=True)
                user_ops_data[key]['cost'] += model_cost
                # DEBUG: Log after adding
                if model and 'flux-2-flex' in model.lower():
                    import sys
                    print(f"DEBUG user_ops AFTER: model={repr(model)}, cost_after={user_ops_data[key]['cost']}", file=sys.stderr, flush=True)
                if status == "charged":
                    price_rubles = float(price) / 100.0
                    user_ops_data[key]['revenue'] += price_rubles
        
        # Sort by telegram_id, then by count descending
        sorted_user_ops = sorted(user_ops_data.items(), key=lambda x: (x[0][0], -x[1]['count']))
        
        for (tg_id, username, first_name, op_type, model), data in sorted_user_ops:
            profit = data['revenue'] - data['cost']
            ws_user_ops.append([
                tg_id,
                f"@{username}" if username else "",
                first_name or "",
                get_operation_name(op_type),
                model or "",
                data['count'],
                data['revenue'],
                data['cost'],
                profit
            ])
        
        # 7. Daily statistics sheet
        ws_daily = wb.create_sheet("Статистика по дням")
        ws_daily.append(["Дата", "Количество операций", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)", "Уникальных пользователей"])
        for cell in ws_daily[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all operations and aggregate by date in Python to handle date-based conversion
        all_daily_ops = db.query(
            Operation.created_at,
            Operation.price,
            Operation.status,
            Operation.user_id,
            Operation.model,
            Operation.type
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        
        daily_data = defaultdict(lambda: {'count': 0, 'revenue': 0.0, 'cost': 0.0, 'users': set()})
        for op in all_daily_ops:
            # Only count operations after migration date
            is_after_migration = False
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif op.price > 100:
                is_after_migration = True
            
            if is_after_migration and op.created_at:
                # Convert to Moscow time before getting date
                moscow_dt = to_moscow_time(op.created_at)
                date_key = moscow_dt.date() if hasattr(moscow_dt, 'date') else moscow_dt
                daily_data[date_key]['count'] += 1
                daily_data[date_key]['users'].add(op.user_id)
                # Calculate cost for all operations (charged and free) - cost is real expense
                # Если модель не указана, определяем по типу операции
                model_cost = get_model_cost_rub(op.model, op.type)
                daily_data[date_key]['cost'] += model_cost
                if op.status == "charged":
                    price_rubles = float(op.price) / 100.0
                    daily_data[date_key]['revenue'] += price_rubles
        
        # Sort by date descending
        sorted_daily = sorted(daily_data.items(), key=lambda x: x[0], reverse=True)
        
        for date, data in sorted_daily:
            profit = data['revenue'] - data['cost']
            ws_daily.append([
                date.strftime("%d.%m.%Y") if isinstance(date, datetime) or hasattr(date, 'strftime') else str(date),
                data['count'],
                data['revenue'],
                data['cost'],
                profit,
                len(data['users'])
            ])
        
        # 8. Weekly statistics sheet
        ws_weekly = wb.create_sheet("Статистика по неделям")
        ws_weekly.append(["Неделя", "Количество операций", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)", "Уникальных пользователей"])
        for cell in ws_weekly[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get week start dates (Monday)
        weekly_data = defaultdict(lambda: {'count': 0, 'revenue': 0.0, 'cost': 0.0, 'users': set()})
        
        operations = db.query(
            Operation.created_at,
            Operation.price,
            Operation.status,
            Operation.user_id,
            Operation.model,
            Operation.type
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        
        for op in operations:
            # Only count operations after migration date
            is_after_migration = False
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif op.price > 100:
                is_after_migration = True
            
            if is_after_migration and op.created_at:
                # Convert to Moscow time first
                moscow_dt = to_moscow_time(op.created_at)
                # Get Monday of the week in Moscow time
                week_start = moscow_dt - timedelta(days=moscow_dt.weekday())
                week_key = week_start.strftime("%d.%m.%Y")
                
                weekly_data[week_key]['count'] += 1
                weekly_data[week_key]['users'].add(op.user_id)
                # Calculate cost for all operations (charged and free) - cost is real expense
                # Если модель не указана, определяем по типу операции
                model_cost = get_model_cost_rub(op.model, op.type)
                weekly_data[week_key]['cost'] += model_cost
                if op.status == "charged":
                    price_rubles = float(op.price) / 100.0
                    weekly_data[week_key]['revenue'] += price_rubles
        
        # Sort by date descending
        sorted_weeks = sorted(weekly_data.items(), key=lambda x: datetime.strptime(x[0], "%d.%m.%Y"), reverse=True)
        
        for week_key, data in sorted_weeks:
            # Calculate week end (Sunday)
            week_start = datetime.strptime(week_key, "%d.%m.%Y")
            week_end = week_start + timedelta(days=6)
            profit = data['revenue'] - data['cost']
            ws_weekly.append([
                f"{week_key} - {week_end.strftime('%d.%m.%Y')}",
                data['count'],
                data['revenue'],
                data['cost'],
                profit,
                len(data['users'])
            ])
        
        # 9. Monthly statistics sheet
        ws_monthly = wb.create_sheet("Статистика по месяцам")
        ws_monthly.append(["Месяц", "Количество операций", "Выручка (₽)", "Себестоимость (₽)", "Прибыль (₽)", "Уникальных пользователей"])
        for cell in ws_monthly[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all operations and aggregate by month in Python to handle date-based conversion
        all_monthly_ops = db.query(
            Operation.created_at,
            Operation.price,
            Operation.status,
            Operation.user_id,
            Operation.model,
            Operation.type
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        
        monthly_data = defaultdict(lambda: {'count': 0, 'revenue': 0.0, 'cost': 0.0, 'users': set()})
        for op in all_monthly_ops:
            # Only count operations after migration date
            is_after_migration = False
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    is_after_migration = True
            elif op.price > 100:
                is_after_migration = True
            
            if is_after_migration and op.created_at:
                # Convert to Moscow time first
                moscow_dt = to_moscow_time(op.created_at)
                # Get year and month from Moscow time
                if hasattr(moscow_dt, 'year'):
                    year = moscow_dt.year
                    month = moscow_dt.month
                else:
                    # Fallback for string dates
                    year = int(str(moscow_dt)[:4])
                    month = int(str(moscow_dt)[5:7])
                
                month_key = (year, month)
                monthly_data[month_key]['count'] += 1
                monthly_data[month_key]['users'].add(op.user_id)
                # Calculate cost for all operations (charged and free) - cost is real expense
                # Если модель не указана, определяем по типу операции
                model_cost = get_model_cost_rub(op.model, op.type)
                monthly_data[month_key]['cost'] += model_cost
                if op.status == "charged":
                    price_rubles = float(op.price) / 100.0
                    monthly_data[month_key]['revenue'] += price_rubles
        
        # Russian month names
        month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
        }
        
        # Sort by year and month descending
        sorted_monthly = sorted(monthly_data.items(), key=lambda x: (x[0][0], x[0][1]), reverse=True)
        
        for (year, month), data in sorted_monthly:
            month_name = f"{month_names.get(month, str(month))} {year}"
            profit = data['revenue'] - data['cost']
            ws_monthly.append([
                month_name,
                data['count'],
                data['revenue'],
                data['cost'],
                profit,
                len(data['users'])
            ])
        
        # 10. AI Assistant Questions sheet
        ws_ai_questions = wb.create_sheet("Вопросы ИИ-помощнику")
        ws_ai_questions.append([
            "ID", "Telegram ID", "Username", "Имя", "Дата и время", "Вопрос", "Ответ", "Ошибка"
        ])
        for cell in ws_ai_questions[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all AI assistant questions with user info
        try:
            # First, check total count without join
            total_questions = db.query(AiAssistantQuestion).count()
            print(f"Total AI assistant questions in DB: {total_questions}")
            
            ai_questions = db.query(
                AiAssistantQuestion,
                User.telegram_id,
                User.username,
                User.first_name
            ).join(
                User, AiAssistantQuestion.user_id == User.id
            ).order_by(
                desc(AiAssistantQuestion.created_at)
            ).all()
            
            print(f"Found {len(ai_questions)} AI assistant questions (with user join)")
            
            # If join returns 0 but total > 0, there might be orphaned questions
            if total_questions > 0 and len(ai_questions) == 0:
                print(f"⚠️ Warning: {total_questions} questions exist but join returned 0. Checking for orphaned questions...")
                orphaned = db.query(AiAssistantQuestion).filter(
                    ~AiAssistantQuestion.user_id.in_(db.query(User.id))
                ).count()
                print(f"Orphaned questions (without matching users): {orphaned}")
            
            rows_added = 0
            for question_obj, tg_id, username, first_name in ai_questions:
                # Format created_at in Moscow time
                created_at_str = format_datetime_moscow(question_obj.created_at, "%d.%m.%Y %H:%M:%S") if question_obj.created_at and not isinstance(question_obj.created_at, str) else (question_obj.created_at if question_obj.created_at else "")
                
                # Truncate long text fields for Excel
                question_text = question_obj.question[:500] if question_obj.question else ""
                answer_text = question_obj.answer[:1000] if question_obj.answer else ""
                error_text = question_obj.error[:500] if question_obj.error else ""
                
                ws_ai_questions.append([
                    question_obj.id,
                    tg_id,
                    username or "",
                    first_name or "",
                    created_at_str,
                    question_text,
                    answer_text,
                    error_text
                ])
                rows_added += 1
                print(f"  Added row {rows_added}: ID={question_obj.id}, Question={question_text[:30]}...")
            
            print(f"Total rows added to Excel sheet: {rows_added}")
        except Exception as e:
            print(f"Error exporting AI assistant questions: {e}")
            import traceback
            traceback.print_exc()
            # Add error row
            ws_ai_questions.append([
                "Ошибка",
                "",
                "",
                "",
                "",
                f"Ошибка при экспорте: {str(e)}",
                "",
                ""
            ])
        
        wb.save(output_path)
        print(f"✅ Статистика экспортирована в файл: {output_path.resolve()}")
        return str(output_path.resolve())
        
    except Exception as e:
        print(f"❌ Ошибка при экспорте: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        db.close()


if __name__ == "__main__":
    output_file = sys.argv[1] if len(sys.argv) > 1 else "statistics_export.xlsx"
    export_statistics_to_excel(output_file)
