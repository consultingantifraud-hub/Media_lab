#!/usr/bin/env python3
"""Export user operations to Excel file."""
import sys
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Optional

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.db.base import SessionLocal
from app.db.models import User, Operation, Payment, PaymentStatus
from app.services.pricing import get_operation_name
from sqlalchemy import desc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Moscow timezone (UTC+3)
MOSCOW_TZ = timezone(timedelta(hours=3))

def to_moscow_time(dt: datetime | None) -> datetime | None:
    """Convert datetime to Moscow timezone (UTC+3)."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(MOSCOW_TZ)

def format_datetime_moscow(dt: datetime | None, format_str: str = "%d.%m.%Y %H:%M") -> str:
    """Format datetime in Moscow timezone."""
    if dt is None:
        return ""
    moscow_dt = to_moscow_time(dt)
    return moscow_dt.strftime(format_str)

def export_user_operations_to_excel(
    user_id: int,
    days: Optional[int],
    output_file: str = "user_operations.xlsx"
) -> Optional[str]:
    """
    Export user operations to Excel file.
    
    Args:
        user_id: User ID
        days: Number of days to filter (None = all)
        output_file: Output file path
        
    Returns:
        Path to created file or None on error
    """
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            print(f"User {user_id} not found")
            return None
        
        # Calculate date filter if needed
        date_filter = None
        if days:
            date_filter = datetime.now(timezone.utc) - timedelta(days=days)
        
        # Get operations
        operations_query = db.query(Operation).filter(Operation.user_id == user_id)
        if date_filter:
            operations_query = operations_query.filter(Operation.created_at >= date_filter)
        operations = operations_query.order_by(desc(Operation.created_at)).all()
        
        # Get payments
        payments_query = db.query(Payment).filter(
            Payment.user_id == user_id,
            Payment.status == PaymentStatus.SUCCEEDED
        )
        if date_filter:
            payments_query = payments_query.filter(Payment.created_at >= date_filter)
        payments = payments_query.order_by(desc(Payment.created_at)).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Операции"
        
        # Headers
        headers = [
            "Дата и время",
            "Тип операции",
            "Статус",
            "Стоимость (₽)",
            "Оригинальная стоимость (₽)",
            "Скидка (%)",
            "Сумма скидки (₽)",
            "ID операции"
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Operation type mapping
        type_names = {
            "generate": "Генерация изображения",
            "edit": "Изменить",
            "merge": "Объединение",
            "retouch": "Ретушь",
            "upscale": "Улучшить",
            "prompt_generation": "Генерация промпта",
            "face_swap": "Заменить лицо",
            "add_text": "Добавить текст",
            "payment": "Пополнение баланса",
        }
        
        # Status mapping
        status_names = {
            "charged": "Списано",
            "pending": "Ожидает",
            "failed": "Ошибка",
            "free": "Бесплатно",
            "refunded": "Возврат",
            "succeeded": "Успешно",
        }
        
        # Combine operations and payments
        all_records = []
        
        # Add operations
        for op in operations:
            price_rubles = op.price / 100.0 if op.price else 0.0
            original_price_rubles = op.original_price / 100.0 if op.original_price else None
            discount_percent = op.discount_percent
            discount_amount_rubles = None
            if original_price_rubles and discount_percent:
                discount_amount_rubles = original_price_rubles - price_rubles
            
            all_records.append({
                "created_at": op.created_at,
                "type": type_names.get(op.type, op.type),
                "status": status_names.get(op.status.value, op.status.value),
                "price": price_rubles,
                "original_price": original_price_rubles,
                "discount_percent": discount_percent,
                "discount_amount": discount_amount_rubles,
                "id": op.id,
                "task_id": op.task_id,
            })
        
        # Add payments
        for payment in payments:
            price_rubles = payment.amount / 100.0 if payment.amount else 0.0
            all_records.append({
                "created_at": payment.created_at,
                "type": "Пополнение баланса",
                "status": "Успешно",
                "price": price_rubles,
                "original_price": None,
                "discount_percent": None,
                "discount_amount": None,
                "id": payment.id,
                "task_id": None,
            })
        
        # Sort by created_at descending
        all_records.sort(key=lambda x: x["created_at"], reverse=True)
        
        # Add rows
        for record in all_records:
            ws.append([
                format_datetime_moscow(record["created_at"], "%d.%m.%Y %H:%M:%S"),
                record["type"],
                record["status"],
                record["price"],
                record["original_price"] if record["original_price"] else "",
                record["discount_percent"] if record["discount_percent"] else "",
                record["discount_amount"] if record["discount_amount"] else "",
                record["id"],
            ])
        
        # Auto-adjust column widths
        column_widths = {
            "A": 20,  # Дата и время
            "B": 25,  # Тип операции
            "C": 12,  # Статус
            "D": 15,  # Стоимость
            "E": 20,  # Оригинальная стоимость
            "F": 12,  # Скидка %
            "G": 15,  # Сумма скидки
            "H": 12,  # ID операции
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Format price columns as numbers with 2 decimal places
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=7):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Save file
        output_path = Path(output_file)
        wb.save(output_path)
        print(f"✅ Операции экспортированы в файл: {output_path.resolve()}")
        return str(output_path.resolve())
        
    except Exception as e:
        print(f"❌ Ошибка при экспорте: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_user_operations.py <user_id> <days> [output_file]")
        sys.exit(1)
    
    user_id = int(sys.argv[1])
    days = int(sys.argv[2]) if sys.argv[2] != "None" else None
    output_file = sys.argv[3] if len(sys.argv) > 3 else "user_operations.xlsx"
    
    export_user_operations_to_excel(user_id, days, output_file)

