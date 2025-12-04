#!/usr/bin/env python3
"""Verify Excel export content."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.db.base import SessionLocal
from scripts.export_statistics_to_excel import export_statistics_to_excel
from openpyxl import load_workbook
import tempfile
import os

# Create test export
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    test_file = tmp.name

try:
    print("Создание тестового экспорта...")
    export_statistics_to_excel(test_file)
    
    print(f"\nЧтение файла: {test_file}")
    wb = load_workbook(test_file)
    
    print("\n" + "=" * 80)
    print("ПРОВЕРКА СОДЕРЖИМОГО ЛИСТОВ")
    print("=" * 80)
    
    # Check each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📊 Лист: {sheet_name}")
        print(f"   Строк: {ws.max_row}, Колонок: {ws.max_column}")
        
        # Show first few rows
        max_rows_to_show = min(10, ws.max_row)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows_to_show, values_only=True), 1):
            row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
            print(f"   {row_idx}: {row_str}")
        
        # Special checks for key sheets
        if sheet_name == "Сводка":
            print("\n   ✅ Проверка листа 'Сводка':")
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if row[0] and isinstance(row[0], str) and "заработано" in row[0].lower():
                    print(f"      {row[0]}: {row[1]} (тип: {type(row[1]).__name__})")
        
        elif sheet_name == "Операции по типам":
            print("\n   ✅ Проверка листа 'Операции по типам':")
            total_revenue = 0.0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row and len(row) >= 3 and isinstance(row[2], (int, float)):
                    revenue = float(row[2])
                    total_revenue += revenue
                    print(f"      {row[0]}: {row[1]} операций, {revenue:.2f} ₽")
            print(f"      ИТОГО: {total_revenue:.2f} ₽")
        
        elif sheet_name == "Пользователи":
            print("\n   ✅ Проверка листа 'Пользователи':")
            total_spent = 0.0
            for row in ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True):
                if row and len(row) >= 12:
                    spent = row[11] if isinstance(row[11], (int, float)) else 0.0
                    total_spent += float(spent)
                    print(f"      Пользователь {row[1]}: {row[10]} операций, {spent:.2f} ₽")
            print(f"      ИТОГО потрачено: {total_spent:.2f} ₽")
    
    print("\n" + "=" * 80)
    print("✅ ПРОВЕРКА ЗАВЕРШЕНА")
    print("=" * 80)
    
except Exception as e:
    print(f"\n❌ Ошибка: {e}")
    import traceback
    traceback.print_exc()
finally:
    if os.path.exists(test_file):
        os.remove(test_file)





