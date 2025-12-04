#!/usr/bin/env python3
"""Test export of AI assistant questions."""
import sys
from pathlib import Path
import tempfile
import os

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.export_statistics_to_excel import export_statistics_to_excel

if __name__ == "__main__":
    # Create temporary file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        excel_file = tmp.name
    
    try:
        print("Running export...")
        result = export_statistics_to_excel(excel_file)
        print(f"Export result: {result}")
        
        # Check if file exists and has content
        if os.path.exists(excel_file):
            file_size = os.path.getsize(excel_file)
            print(f"File created: {excel_file}, size: {file_size} bytes")
            
            # Try to read the file and check for AI questions sheet
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            print(f"\nSheets in workbook: {wb.sheetnames}")
            
            if "Вопросы ИИ-помощнику" in wb.sheetnames:
                ws = wb["Вопросы ИИ-помощнику"]
                print(f"\nSheet 'Вопросы ИИ-помощнику' found!")
                print(f"Max row: {ws.max_row}")
                print(f"Max column: {ws.max_column}")
                
                if ws.max_row > 1:
                    print("\nFirst few rows:")
                    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                        if i <= 5:
                            print(f"Row {i}: {row}")
                else:
                    print("⚠️ Sheet has only header row, no data!")
            else:
                print("❌ Sheet 'Вопросы ИИ-помощнику' not found!")
        else:
            print(f"❌ File not created: {excel_file}")
    finally:
        # Clean up
        if os.path.exists(excel_file):
            os.unlink(excel_file)




