"""Billing handlers for Telegram bot."""
import asyncio
from aiogram import Router, F
from aiogram.filters import Command, StateFilter
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from loguru import logger
from app.billing import build_operations_history_keyboard

from app.services.billing import BillingService, get_user_info
from app.services.payment import PaymentService, create_payment
from app.services.discount import DiscountService
from app.services.pricing import get_all_prices, get_operation_name
from app.db.base import SessionLocal
from app.bot.keyboards.main import BALANCE_BUTTON
from app.db.models import OperationStatus, User, UserStatistics, Operation, Balance
from sqlalchemy import func, desc
from typing import Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import json
try:
    from zoneinfo import ZoneInfo
except ImportError:
    # Fallback for Python < 3.9
    from backports.zoneinfo import ZoneInfo

# Moscow timezone (UTC+3)
MOSCOW_TZ = timezone(timedelta(hours=3))

def get_moscow_time() -> datetime:
    """Get current time in Moscow timezone (UTC+3)."""
    return datetime.now(timezone.utc).astimezone(MOSCOW_TZ)


def format_balance(balance: float | int) -> str:
    """Форматирует баланс с округлением до 2 знаков после запятой (копеек)."""
    return f"{round(float(balance), 2):.2f}"

router = Router()

# Fixed payment amounts
PAYMENT_AMOUNTS = [100, 300, 500, 1000]


class PaymentStates(StatesGroup):
    """States for payment flow."""
    WAIT_CUSTOM_AMOUNT = State()
    WAIT_DISCOUNT_CODE = State()  # For payment discount codes
    WAIT_EMAIL = State()  # For email for receipt
    BALANCE_MENU_SHOWN = State()  # Balance menu was shown (to intercept text input)


class OperationDiscountStates(StatesGroup):
    """States for operation discount codes."""
    WAIT_OPERATION_DISCOUNT_CODE = State()


def build_payment_keyboard() -> InlineKeyboardMarkup:
    """Build payment amount selection keyboard."""
    buttons = []
    for amount in PAYMENT_AMOUNTS:
        buttons.append([
            InlineKeyboardButton(
                text=f"{amount} ₽",
                callback_data=f"payment_amount_{amount}"
            )
        ])
    buttons.append([
        InlineKeyboardButton(
            text="🔢 Другая сумма",
            callback_data="payment_custom"
        )
    ])
    buttons.append([
        InlineKeyboardButton(
            text="🎟️ Ввести промокод",
            callback_data="payment_discount_code"
        )
    ])
    buttons.append([
        InlineKeyboardButton(
            text="❌ Отмена",
            callback_data="payment_cancel"
        )
    ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def build_balance_keyboard() -> InlineKeyboardMarkup:
    """Build balance menu keyboard."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text="💰 Пополнить баланс",
                callback_data="payment_menu"
            )
        ],
        [
            InlineKeyboardButton(
                text="🎟️ Промокод для операций",
                callback_data="operation_discount_code"
            )
        ],
        [
            InlineKeyboardButton(
                text="📊 История операций",
                callback_data="operations_history"
            )
        ],
        [
            InlineKeyboardButton(
                text="❌ Закрыть",
                callback_data="balance_close"
            )
        ]
    ])


def log_history_keyboard(callback: CallbackQuery, keyboard: InlineKeyboardMarkup, source: str = "unknown") -> None:
    logger.info(
        "HISTORY KB ROWS (from %s): %s",
        callback.data if hasattr(callback, "data") else source,
        [[btn.text for btn in row] for row in keyboard.inline_keyboard],
    )


async def check_last_payment(message: Message):
    """Check status of last payment and update balance if needed."""
    from app.services.payment import PaymentService
    from app.db.models import Payment, PaymentStatus
    from datetime import datetime, timedelta
    
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
        
        # Find last pending payment for this user
        last_payment = db.query(Payment).filter(
            Payment.user_id == user.id,
            Payment.status.in_([PaymentStatus.PENDING, PaymentStatus.STALE])
        ).order_by(Payment.created_at.desc()).first()
        
        if not last_payment:
            await message.answer("ℹ️ У вас нет активных платежей.")
            return
        
        if not last_payment.yookassa_payment_id:
            await message.answer("⚠️ Платеж найден, но ID YooKassa отсутствует.")
            return
        
        if last_payment.status == PaymentStatus.STALE:
            await message.answer(
                "ℹ️ Платёж всё ещё подтверждается YooKassa.\n"
                "Мы автоматически повторяем проверки, но вы можете запустить ручную проверку ниже.",
                parse_mode="Markdown",
            )
        
        # Check status from YooKassa
        await message.answer("⏳ Проверяю статус платежа...")
        
        status_info = PaymentService.check_payment_status_from_yookassa(
            db, last_payment.yookassa_payment_id
        )
        
        if status_info:
            if status_info["status"] == "succeeded" and status_info["paid"]:
                balance_after = BillingService.get_user_balance(db, user.id)
                balance_after_rubles = balance_after / 100.0
                if status_info.get("credited", False):
                    text = (
                        "✅ **Оплата успешно подтверждена!**\n\n"
                        f"💰 Ваш баланс пополнен на {status_info['amount']:.2f}₽\n"
                        f"💵 Текущий баланс: {balance_after_rubles:.2f}₽"
                    )
                else:
                    text = (
                        "ℹ️ **Оплата уже была зачислена ранее.**\n\n"
                        f"💵 Текущий баланс: {balance_after_rubles:.2f}₽"
                    )
                await message.answer(text, parse_mode="Markdown")
            elif status_info["status"] == "pending":
                await message.answer(
                    "⏳ **Платеж обрабатывается...**\n\n"
                    "Пожалуйста, подождите несколько секунд и попробуйте снова.",
                    parse_mode="Markdown"
                )
            else:
                await message.answer(
                    f"❌ **Платеж не завершен**\n\n"
                    f"Статус: {status_info['status']}\n"
                    "Если вы оплатили, но баланс не пополнился, обратитесь в поддержку.",
                    parse_mode="Markdown"
                )
        else:
            await message.answer(
                "⚠️ **Не удалось проверить статус платежа**\n\n"
                "Попробуйте позже или обратитесь в поддержку.",
                parse_mode="Markdown"
            )
    except Exception as e:
        logger.error(f"Error checking payment status: {e}", exc_info=True)
        await message.answer("❌ Произошла ошибка при проверке платежа.")
    finally:
        db.close()


async def show_balance(message: Message, state: FSMContext = None):
    """Show user balance with prices."""
    # Set state to indicate balance menu was shown
    if state:
        await state.set_state(PaymentStates.BALANCE_MENU_SHOWN)
        logger.info("Balance menu shown, state set to BALANCE_MENU_SHOWN for user_id={}", 
                   message.from_user.id if message.from_user else None)
    
    db = SessionLocal()
    try:
        # Check for pending payments and update if needed (async, non-blocking)
        # ВАЖНО: Не блокируем обработку баланса проверкой платежей
        # Проверка платежей может занимать до 60+ секунд из-за timeout и retry
        # Выполняем проверку в фоне через asyncio.create_task
        from app.services.payment import PaymentService
        from app.db.models import Payment, PaymentStatus, User
        user_obj = db.query(User).filter(User.telegram_id == message.from_user.id).first()
        if user_obj:
            pending_payments = db.query(Payment).filter(
                Payment.user_id == user_obj.id,
                Payment.status.in_([PaymentStatus.PENDING, PaymentStatus.STALE])
            ).order_by(Payment.created_at.desc()).limit(1).all()
            
            # Запускаем проверку платежей в фоне, не блокируя ответ пользователю
            if pending_payments:
                async def check_payments_background():
                    """Check payment status in background without blocking."""
                    db_bg = SessionLocal()
                    try:
                        for payment in pending_payments:
                            if payment.yookassa_payment_id:
                                # Check status from YooKassa (silently, don't show errors to user)
                                try:
                                    # Используем run_in_executor для синхронной функции
                                    loop = asyncio.get_event_loop()
                                    await loop.run_in_executor(
                                        None,
                                        PaymentService.check_payment_status_from_yookassa,
                                        db_bg, payment.yookassa_payment_id
                                    )
                                except Exception as e:
                                    logger.debug(f"Error checking payment status in background: {e}")
                    finally:
                        db_bg.close()
                
                # Запускаем проверку в фоне, не ждем результата
                asyncio.create_task(check_payments_background())
        
        user_info = BillingService.get_user_info(db, message.from_user.id)
        if not user_info:
            # Create user if doesn't exist
            user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
            user_info = BillingService.get_user_info(db, message.from_user.id)

        balance = user_info["balance"]
        has_free_access = user_info.get("has_free_access", False)

        # Get user object to check for active discount code
        from app.db.models import User, DiscountCode
        user = db.query(User).filter(User.telegram_id == message.from_user.id).first()
        
        # Check for active operation discount code
        discount_info = ""
        if user and user.operation_discount_code_id and user.operation_discount_percent:
            discount_code = db.query(DiscountCode).filter(DiscountCode.id == user.operation_discount_code_id).first()
            if discount_code:
                discount_info = (
                    f"\n\n🎟️ **Активный промокод:** {discount_code.code}\n"
                    f"💰 **Скидка на операции:** {user.operation_discount_percent}%"
                )

        # Get prices for display (already sorted by price in descending order)
        prices = get_all_prices()
        
        # Отладочное логирование для проверки Flux 2 Flex
        logger.debug(f"show_balance: All prices keys: {list(prices.keys())}")
        logger.debug(f"show_balance: Flux 2 Flex in prices: {'Flux 2 Flex (генерация)' in prices}")
        if "Flux 2 Flex (генерация)" in prices:
            logger.debug(f"show_balance: Flux 2 Flex price: {prices['Flux 2 Flex (генерация)']}")
        
        # Формируем список услуг с ценами (уже отсортирован по убыванию)
        services_list = []
        for service_name, price in prices.items():
            # Упрощаем названия для отображения
            if service_name == "Nano Banana Pro (генерация/объединение)":
                services_list.append(f"• Nano Banana Pro: {price} ₽")
            elif service_name == "Flux 2 Flex (генерация)":
                services_list.append(f"• Flux 2 Flex: {price} ₽")
            elif service_name == "Seedream (генерация/редактирование)":
                services_list.append(f"• Seedream: {price} ₽")
            elif service_name == "Nano Banana (генерация/редактирование)":
                services_list.append(f"• Nano Banana: {price} ₽")
            elif service_name == "Остальные модели (генерация/редактирование/объединение/ретушь/upscale)":
                services_list.append(f"• Ретушь, Улучшить: {price} ₽")
            elif service_name == "Генерация промпта":
                services_list.append(f"• Генерация промпта: {price} ₽")
            elif service_name == "Замена лица":
                services_list.append(f"• Замена лица: {price} ₽")
            elif service_name == "Добавление текста":
                services_list.append(f"• Добавление текста: {price} ₽")
            else:
                # Добавляем все остальные услуги, которые не были обработаны
                services_list.append(f"• {service_name}: {price} ₽")
        
        services_text = "\n".join(services_list)
        
        if has_free_access:
            text = (
                f"💰 **Ваш баланс:** {format_balance(balance)} ₽\n\n"
                f"✨ **Бесплатный доступ:** Активен\n"
                f"💡 Вы можете пользоваться сервисом бесплатно без ограничений\n\n"
                f"📋 **Базовая стоимость услуг (без скидки):**\n"
                f"{services_text}"
                f"{discount_info}"
            )
        else:
            text = (
                f"💰 **Ваш баланс:** {format_balance(balance)} ₽"
                f"{discount_info}\n\n"
                f"📋 **Базовая стоимость услуг (без скидки):**\n"
                f"{services_text}"
            )

        await message.answer(
            text,
            reply_markup=build_balance_keyboard(),
            parse_mode="Markdown"
        )
    finally:
        db.close()


@router.message(Command("check_payment"))
async def cmd_check_payment(message: Message):
    """Check status of last payment."""
    await check_last_payment(message)


@router.message(Command("balance"))
async def cmd_balance(message: Message, state: FSMContext):
    """Show user balance (command handler)."""
    await show_balance(message, state)


@router.message(Command("add_balance"))
async def handle_add_balance(message: Message):
    """
    Command to add balance directly.
    Usage: /add_balance <amount>
    Example: /add_balance 500
    """
    db = SessionLocal()
    try:
        # Parse amount from command
        parts = message.text.split()
        if len(parts) < 2:
            await message.answer(
                "❌ Укажите сумму для пополнения.\n\n"
                "Пример: `/add_balance 500`",
                parse_mode="Markdown"
            )
            return
        
        try:
            amount = int(parts[1])
            if amount <= 0:
                await message.answer("❌ Сумма должна быть больше 0.")
                return
            if amount > 100000:
                await message.answer("❌ Максимальная сумма: 100,000 ₽")
                return
        except ValueError:
            await message.answer("❌ Неверный формат суммы. Используйте целое число.")
            return
        
        # Get or create user
        user, created = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
        
        # Add balance
        success = BillingService.add_balance(db, user.id, amount)
        
        if success:
            # Get updated balance
            user_info = BillingService.get_user_info(db, message.from_user.id)
            new_balance = user_info["balance"] if user_info else 0
            
            await message.answer(
                f"✅ **Баланс пополнен**\n\n"
                f"💰 Добавлено: {amount} ₽\n"
                f"💵 Новый баланс: {format_balance(new_balance)} ₽",
                parse_mode="Markdown"
            )
            logger.info(f"Balance added: user_id={user.id}, telegram_id={message.from_user.id}, amount={amount}₽, new_balance={new_balance}₽")
        else:
            await message.answer("❌ Ошибка при пополнении баланса.")
            
    except Exception as e:
        logger.error(f"Error in add_balance: {e}", exc_info=True)
        await message.answer(f"❌ Ошибка: {str(e)}")
    finally:
        db.close()


@router.message(Command("test_add_balance"))
async def handle_test_add_balance(message: Message):
    """
    Test command to add balance directly (for testing without YooKassa).
    Usage: /test_add_balance <amount>
    Example: /test_add_balance 500
    """
    db = SessionLocal()
    try:
        # Parse amount from command
        parts = message.text.split()
        if len(parts) < 2:
            await message.answer(
                "❌ Укажите сумму для пополнения.\n\n"
                "Пример: `/test_add_balance 500`",
                parse_mode="Markdown"
            )
            return
        
        try:
            amount = int(parts[1])
            if amount <= 0:
                await message.answer("❌ Сумма должна быть больше 0.")
                return
            if amount > 100000:
                await message.answer("❌ Максимальная сумма: 100,000 ₽")
                return
        except ValueError:
            await message.answer("❌ Неверный формат суммы. Используйте целое число.")
            return
        
        # Get or create user
        user, created = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
        
        # Add balance
        success = BillingService.add_balance(db, user.id, amount)
        
        if success:
            # Get updated balance
            user_info = BillingService.get_user_info(db, message.from_user.id)
            new_balance = user_info["balance"] if user_info else 0
            
            await message.answer(
                f"✅ **Тестовое пополнение баланса**\n\n"
                f"💰 Добавлено: {amount} ₽\n"
                f"💵 Новый баланс: {new_balance} ₽\n\n"
                f"⚠️ Это тестовая команда для разработки.",
                parse_mode="Markdown"
            )
            logger.info(f"Test balance added: user_id={user.id}, amount={amount}₽, new_balance={new_balance}₽")
        else:
            await message.answer("❌ Ошибка при пополнении баланса.")
            
    except Exception as e:
        logger.error(f"Error in test_add_balance: {e}", exc_info=True)
        await message.answer(f"❌ Ошибка: {str(e)}")
    finally:
        db.close()


# Обработчик баланса регистрируется в register_billing_handlers через dp.message.register
# для обеспечения высокого приоритета над общим обработчиком текста
async def handle_balance_button(message: Message, state: FSMContext):
    """Handle balance button click."""
    logger.info("handle_balance_button called: user_id={}, text='{}'", 
               message.from_user.id if message.from_user else None, 
               message.text)
    await show_balance(message, state)


@router.callback_query(F.data == "payment_menu")
async def callback_payment_menu(callback: CallbackQuery, state: FSMContext):
    """Show payment menu."""
    # Keep BALANCE_MENU_SHOWN state to intercept text input
    # State will be cleared when user selects specific amount or clicks "Другая сумма"
    db = SessionLocal()
    try:
        user_info = BillingService.get_user_info(db, callback.from_user.id)
        if not user_info:
            user, _ = BillingService.get_or_create_user(db, callback.from_user.id, callback.from_user)
            user_info = BillingService.get_user_info(db, callback.from_user.id)

        balance = user_info["balance"]
        has_free_access = user_info.get("has_free_access", False)

        # Get user object to check for active discount code
        from app.db.models import User, DiscountCode
        user = db.query(User).filter(User.telegram_id == callback.from_user.id).first()
        
        # Check for active operation discount code
        discount_info = ""
        if user and user.operation_discount_code_id and user.operation_discount_percent:
            discount_code = db.query(DiscountCode).filter(DiscountCode.id == user.operation_discount_code_id).first()
            if discount_code:
                discount_info = (
                    f"\n🎟️ **Активный промокод:** {discount_code.code}\n"
                    f"💰 **Скидка на операции:** {user.operation_discount_percent}%"
                )

        prices = get_all_prices()
        
        # Формируем список услуг с ценами (уже отсортирован по убыванию)
        services_list = []
        for service_name, price in prices.items():
            # Упрощаем названия для отображения
            if service_name == "Nano Banana Pro (генерация/объединение)":
                services_list.append(f"• Nano Banana Pro: {price} ₽")
            elif service_name == "Flux 2 Flex (генерация)":
                services_list.append(f"• Flux 2 Flex: {price} ₽")
            elif service_name == "Seedream (генерация/редактирование)":
                services_list.append(f"• Seedream: {price} ₽")
            elif service_name == "Nano Banana (генерация/редактирование)":
                services_list.append(f"• Nano Banana: {price} ₽")
            elif service_name == "Остальные модели (генерация/редактирование/объединение/ретушь/upscale)":
                services_list.append(f"• Ретушь, Улучшить: {price} ₽")
            elif service_name == "Генерация промпта":
                services_list.append(f"• Генерация промпта: {price} ₽")
            elif service_name == "Замена лица":
                services_list.append(f"• Замена лица: {price} ₽")
            elif service_name == "Добавление текста":
                services_list.append(f"• Добавление текста: {price} ₽")
            else:
                # Добавляем все остальные услуги, которые не были обработаны
                services_list.append(f"• {service_name}: {price} ₽")
        
        services_text = "\n".join(services_list)
        
        if has_free_access:
            text = (
                f"💰 **Ваш баланс:** {balance} ₽\n"
                f"✨ **Бесплатный доступ:** Активен"
                f"{discount_info}\n\n"
                f"📋 **Базовая стоимость услуг (без скидки):**\n"
                f"{services_text}\n\n"
                f"Выберите сумму для пополнения (опционально):"
            )
        else:
            text = (
                f"💰 **Ваш баланс:** {format_balance(balance)} ₽"
                f"{discount_info}\n\n"
                f"📋 **Базовая стоимость услуг (без скидки):**\n"
                f"{services_text}\n\n"
                f"Выберите сумму для пополнения:"
            )

        await callback.message.edit_text(
            text,
            reply_markup=build_payment_keyboard(),
            parse_mode="Markdown"
        )
        await callback.answer()
    finally:
        db.close()


@router.callback_query(F.data.startswith("payment_amount_"))
async def callback_payment_amount(callback: CallbackQuery, state: FSMContext):
    """Handle fixed amount payment."""
    amount = int(callback.data.split("_")[-1])
    
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, callback.from_user.id)
        
        # Check if user has email (required for receipt)
        if not user.email:
            await state.update_data(payment_amount=amount)
            await state.set_state(PaymentStates.WAIT_EMAIL)
            
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="❌ Отмена",
                        callback_data="payment_cancel"
                    )
                ]
            ])
            
            await callback.message.edit_text(
                "📧 **Введите адрес электронной почты**\n\n"
                "На этот адрес будет отправлен чек об оплате.\n"
                "Введите ваш email:",
                reply_markup=keyboard,
                parse_mode="Markdown"
            )
            await callback.answer()
            return
        
        # Создаем платеж асинхронно, чтобы не блокировать ответ пользователю
        await callback.answer("⏳ Создаю платеж...")
        
        import asyncio
        loop = asyncio.get_event_loop()
        
        try:
            payment_result = await loop.run_in_executor(
                None,
                PaymentService.create_payment,
                db,
                user.id,
                amount,
                f"Пополнение баланса на {amount}₽",
                user.email
            )
        except Exception as e:
            logger.error(f"Error creating payment in background: {e}", exc_info=True)
            await callback.answer("❌ Ошибка при создании платежа. Попробуйте позже.", show_alert=True)
            return

        if not payment_result:
            await callback.answer("❌ Ошибка при создании платежа. Попробуйте позже.", show_alert=True)
            return

        confirmation_url = payment_result["confirmation_url"]
        
        # Create payment link button
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="💳 Оплатить",
                    url=confirmation_url
                )
            ],
            [
                InlineKeyboardButton(
                    text="↩️ Назад",
                    callback_data="payment_menu"
                )
            ]
        ])

        await callback.message.edit_text(
            f"💳 **Платеж создан**\n\n"
            f"Сумма: {amount} ₽\n\n"
            f"Нажмите кнопку ниже для оплаты:",
            reply_markup=keyboard,
            parse_mode="Markdown"
        )
        await callback.answer()
    finally:
        db.close()


@router.callback_query(F.data == "payment_custom")
async def callback_payment_custom(callback: CallbackQuery, state: FSMContext):
    """Request custom payment amount."""
    await state.set_state(PaymentStates.WAIT_CUSTOM_AMOUNT)
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text="❌ Отмена",
                callback_data="payment_cancel"
            )
        ]
    ])

    await callback.message.edit_text(
        "💳 **Введите сумму для пополнения**\n\n"
        "Минимальная сумма: 10 ₽\n"
        "Введите целое число:",
        reply_markup=keyboard,
        parse_mode="Markdown"
    )
    await callback.answer()


async def process_custom_amount(message: Message, state: FSMContext):
    """Process custom payment amount."""
    try:
        amount = int(message.text.strip())
        
        if amount < 10:
            await message.answer("❌ Минимальная сумма пополнения: 10 ₽")
            return

        # Check for discount code in state
        state_data = await state.get_data()
        discount_code = state_data.get("discount_code")

        db = SessionLocal()
        try:
            user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
            
            # Check if user has email (required for receipt)
            if not user.email:
                # Save amount and discount info in state, then ask for email
                await state.update_data(
                    payment_amount=amount,
                    discount_code=discount_code
                )
                await state.set_state(PaymentStates.WAIT_EMAIL)
                
                keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="❌ Отмена",
                            callback_data="payment_cancel"
                        )
                    ]
                ])
                
                await message.answer(
                    "📧 **Введите адрес электронной почты**\n\n"
                    "На этот адрес будет отправлен чек об оплате.\n"
                    "Введите ваш email:",
                    reply_markup=keyboard,
                    parse_mode="Markdown"
                )
                return
            
            # Apply discount if code exists
            final_amount = amount
            discount_amount = 0
            discount_percent = 0
            
            if discount_code:
                is_valid, discount, error_msg = DiscountService.validate_discount_code(
                    db, discount_code, user.id
                )
                if is_valid:
                    discount_percent = discount.discount_percent
                    discount_amount = int(amount * discount_percent / 100)
                    final_amount = amount - discount_amount
            
            # Создаем платеж асинхронно, чтобы не блокировать ответ пользователю
            # PaymentService.create_payment может занимать до 60+ секунд из-за timeout и retry
            import asyncio
            loop = asyncio.get_event_loop()
            
            # Показываем пользователю, что платеж обрабатывается
            processing_msg = await message.answer("⏳ Создаю платеж...")
            
            try:
                payment_result = await loop.run_in_executor(
                    None,
                    PaymentService.create_payment,
                    db,
                    user.id,
                    final_amount,
                    f"Пополнение баланса на {amount}₽" + (f" (скидка {discount_percent}%)" if discount_amount > 0 else ""),
                    user.email
                )
            except Exception as e:
                logger.error(f"Error creating payment in background: {e}", exc_info=True)
                await processing_msg.delete()
                await message.answer("❌ Ошибка при создании платежа. Попробуйте позже.")
                return
            
            # Удаляем сообщение "Создаю платеж..."
            try:
                await processing_msg.delete()
            except Exception as del_err:
                logger.warning(f"Failed to delete processing message: {del_err}")

            if not payment_result:
                logger.error(f"Payment creation returned None for user_id={user.id}, amount={amount}₽")
                await message.answer("❌ Ошибка при создании платежа. Попробуйте позже.")
                return

            logger.info(f"Payment created successfully: payment_id={payment_result.get('payment_id')}, confirmation_url={payment_result.get('confirmation_url', 'N/A')[:50]}...")

            # Apply discount to payment if code was used
            if discount_code:
                try:
                    is_valid, discount, _ = DiscountService.validate_discount_code(db, discount_code, user.id)
                    if is_valid and not discount.is_free_generation:
                        payment_id = payment_result["payment_id"]
                        DiscountService.apply_discount_to_payment(db, discount, user.id, payment_id)
                        await state.update_data(discount_code=None)  # Clear discount code after use
                except Exception as discount_err:
                    logger.error(f"Error applying discount: {discount_err}", exc_info=True)

            confirmation_url = payment_result.get("confirmation_url")
            if not confirmation_url:
                logger.error(f"No confirmation_url in payment_result: {payment_result}")
                await message.answer("❌ Ошибка: не получена ссылка на оплату. Попробуйте позже.")
                return
            
            payment_text = f"💳 **Платеж создан**\n\n"
            if discount_amount > 0:
                payment_text += (
                    f"💰 Сумма: {amount} ₽\n"
                    f"🎟️ Скидка ({discount_percent}%): -{discount_amount} ₽\n"
                    f"💵 К оплате: {final_amount} ₽\n\n"
                )
            else:
                payment_text += f"Сумма: {amount} ₽\n\n"
            payment_text += "Нажмите кнопку ниже для оплаты:"
            
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="💳 Оплатить",
                        url=confirmation_url
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="💰 Баланс",
                        callback_data="payment_menu"
                    )
                ]
            ])

            try:
                await message.answer(
                    payment_text,
                    reply_markup=keyboard,
                    parse_mode="Markdown"
                )
                logger.info(f"Payment message sent successfully to user_id={message.from_user.id}")
            except Exception as send_err:
                logger.error(f"Error sending payment message: {send_err}", exc_info=True)
                await message.answer(f"❌ Ошибка при отправке сообщения о платеже. Ссылка: {confirmation_url}")
            
            await state.clear()
        finally:
            db.close()

    except ValueError:
        await message.answer("❌ Пожалуйста, введите целое число (например: 500)")


async def process_email(message: Message, state: FSMContext):
    """Process email input for payment receipt."""
    import re
    
    email = message.text.strip()
    
    # Basic email validation
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        await message.answer("❌ Неверный формат email. Пожалуйста, введите корректный адрес электронной почты.")
        return
    
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
        
        # Save email to user
        user.email = email
        db.commit()
        
        # Get payment amount from state
        state_data = await state.get_data()
        amount = state_data.get("payment_amount")
        discount_code = state_data.get("discount_code")
        
        if not amount:
            await message.answer("❌ Ошибка: сумма платежа не найдена. Попробуйте создать платеж заново.")
            await state.clear()
            return
        
        # Apply discount if code exists
        final_amount = amount
        discount_amount = 0
        discount_percent = 0
        
        if discount_code:
            is_valid, discount, error_msg = DiscountService.validate_discount_code(
                db, discount_code, user.id
            )
            if is_valid:
                discount_percent = discount.discount_percent
                discount_amount = int(amount * discount_percent / 100)
                final_amount = amount - discount_amount
        
        # Создаем платеж асинхронно, чтобы не блокировать ответ пользователю
        # Показываем пользователю, что платеж обрабатывается
        processing_msg = await message.answer("⏳ Создаю платеж...")
        
        import asyncio
        loop = asyncio.get_event_loop()
        
        try:
            payment_result = await loop.run_in_executor(
                None,
                PaymentService.create_payment,
                db,
                user.id,
                final_amount,
                f"Пополнение баланса на {amount}₽" + (f" (скидка {discount_percent}%)" if discount_amount > 0 else ""),
                email
            )
        except Exception as e:
            logger.error(f"Error creating payment in background: {e}", exc_info=True)
            await processing_msg.delete()
            await message.answer("❌ Ошибка при создании платежа. Попробуйте позже.")
            await state.clear()
            return
        
        # Удаляем сообщение "Создаю платеж..."
        try:
            await processing_msg.delete()
        except Exception as del_err:
            logger.warning(f"Failed to delete processing message: {del_err}")
        
        if not payment_result:
            logger.error(f"Payment creation returned None for user_id={user.id}, amount={amount}₽")
            await message.answer("❌ Ошибка при создании платежа. Попробуйте позже.")
            await state.clear()
            return

        logger.info(f"Payment created successfully: payment_id={payment_result.get('payment_id')}, confirmation_url={payment_result.get('confirmation_url', 'N/A')[:50]}...")
        
        # Apply discount to payment if code was used
        if discount_code:
            is_valid, discount, _ = DiscountService.validate_discount_code(db, discount_code, user.id)
            if is_valid and not discount.is_free_generation:
                payment_id = payment_result["payment_id"]
                DiscountService.apply_discount_to_payment(db, discount, user.id, payment_id)
                await state.update_data(discount_code=None)
        
        confirmation_url = payment_result.get("confirmation_url")
        if not confirmation_url:
            logger.error(f"No confirmation_url in payment_result: {payment_result}")
            await message.answer("❌ Ошибка: не получена ссылка на оплату. Попробуйте позже.")
            await state.clear()
            return
        
        payment_text = f"💳 **Платеж создан**\n\n"
        if discount_amount > 0:
            payment_text += (
                f"💰 Сумма: {amount} ₽\n"
                f"🎟️ Скидка ({discount_percent}%): -{discount_amount} ₽\n"
                f"💵 К оплате: {final_amount} ₽\n\n"
            )
        else:
            payment_text += f"Сумма: {amount} ₽\n\n"
        
        payment_text += f"📧 Чек будет отправлен на: {email}\n\n"
        payment_text += "Нажмите кнопку ниже для оплаты:"
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="💳 Оплатить",
                    url=confirmation_url
                )
            ],
            [
                InlineKeyboardButton(
                    text="↩️ Назад",
                    callback_data="payment_menu"
                )
            ]
        ])
        
        try:
            await message.answer(
                payment_text,
                reply_markup=keyboard,
                parse_mode="Markdown"
            )
            logger.info(f"Payment message sent successfully to user_id={message.from_user.id}")
        except Exception as send_err:
            logger.error(f"Error sending payment message: {send_err}", exc_info=True)
            await message.answer(f"❌ Ошибка при отправке сообщения о платеже. Ссылка: {confirmation_url}")
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error processing email: {e}", exc_info=True)
        await message.answer("❌ Произошла ошибка. Попробуйте позже.")
        await state.clear()
    finally:
        db.close()


@router.callback_query(F.data == "payment_cancel")
async def callback_payment_cancel(callback: CallbackQuery, state: FSMContext):
    """Cancel payment."""
    await state.clear()
    await callback.message.delete()
    await callback.answer("Отменено")


@router.callback_query(F.data == "balance_close")
async def callback_balance_close(callback: CallbackQuery, state: FSMContext):
    """Close balance menu."""
    await state.clear()  # Clear state when closing menu
    await callback.message.delete()
    await callback.answer()


@router.callback_query(F.data == "balance_menu")
async def callback_balance_menu(callback: CallbackQuery, state: FSMContext):
    """Show balance menu."""
    await callback.answer()  # Answer callback first to prevent timeout
    await show_balance(callback.message, state)


async def export_operations_to_excel(callback: CallbackQuery, days: int) -> None:
    """Export user operations to Excel file for specified period."""
    try:
        import tempfile
        import os
        from aiogram.types import FSInputFile
        
        logger.info(f"Starting export_operations_to_excel for {days} days")
        
        # Import here to catch import errors
        try:
            from scripts.export_user_operations import export_user_operations_to_excel
        except ImportError as e:
            logger.error(f"Failed to import export_user_operations: {e}", exc_info=True)
            await callback.message.answer("❌ Ошибка: модуль экспорта не найден.")
            await callback.answer()
            return
        
        db = SessionLocal()
        try:
            user, _ = BillingService.get_or_create_user(db, callback.from_user.id)
            logger.info(f"User found: {user.id}, exporting operations for {days} days")
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                excel_file = tmp.name
            
            logger.info(f"Temporary file created: {excel_file}")
            
            try:
                # Export operations
                logger.info(f"Calling export_user_operations_to_excel(user_id={user.id}, days={days}, file={excel_file})")
                result = export_user_operations_to_excel(user.id, days, excel_file)
                
                logger.info(f"Export result: {result}")
                
                if result and os.path.exists(excel_file):
                    # Send file to user
                    period_text = f"{days} дней" if days > 1 else f"{days} день"
                    file = FSInputFile(excel_file, filename=f"operations_{days}days.xlsx")
                    logger.info(f"Sending file to user: {excel_file}")
                    await callback.message.answer_document(
                        document=file,
                        caption=f"📊 История операций за {period_text}\n{get_moscow_time().strftime('%d.%m.%Y %H:%M')}"
                    )
                    logger.info("File sent successfully")
                else:
                    logger.error(f"Export failed or file not created. Result: {result}, File exists: {os.path.exists(excel_file) if excel_file else False}")
                    await callback.message.answer("❌ Произошла ошибка при формировании выгрузки.")
                    await callback.answer()
            finally:
                # Clean up temporary file
                if os.path.exists(excel_file):
                    os.unlink(excel_file)
                    logger.info(f"Temporary file deleted: {excel_file}")
        finally:
            db.close()
    except Exception as e:
        logger.error(f"Error in export_operations_to_excel: {e}", exc_info=True)
        try:
            await callback.message.answer("❌ Произошла ошибка при формировании выгрузки.")
            await callback.answer()
        except:
            pass


@router.callback_query(F.data == "operations_history")
async def callback_operations_history(callback: CallbackQuery, state: FSMContext):
    keyboard = build_operations_history_keyboard()
    log_history_keyboard(callback, keyboard, source="operations_history")
    await callback.message.edit_text(
        "📊 Выберите период для просмотра истории операций:",
        reply_markup=keyboard,
        parse_mode="Markdown",
    )
    await callback.answer()


@router.callback_query(F.data == "operations_back")
async def callback_operations_back(callback: CallbackQuery, state: FSMContext):
    """Return to payment menu from operations history."""
    await callback_payment_menu(callback, state)


@router.callback_query(F.data.startswith("operations_history_"))
async def callback_operations_history_with_filter(callback: CallbackQuery, state: FSMContext, days: Optional[int] = None):
    data = callback.data
    if data == "operations_history_1":
        days = 1
    elif data == "operations_history_7":
        logger.info(f"Exporting operations for 7 days for user {callback.from_user.id}")
        try:
            await callback.answer("📊 Формирую выгрузку за 7 дней...")
            await export_operations_to_excel(callback, 7)
        except Exception as e:
            logger.error(f"Error exporting operations for 7 days: {e}", exc_info=True)
            await callback.message.answer("❌ Произошла ошибка при формировании выгрузки.")
            await callback.answer()
        keyboard = build_operations_history_keyboard()
        log_history_keyboard(callback, keyboard, source="operations_history_7")
        await callback.message.edit_reply_markup(reply_markup=keyboard)
        return
    elif data == "operations_history_30":
        logger.info(f"Exporting operations for 30 days for user {callback.from_user.id}")
        try:
            await callback.answer("📊 Формирую выгрузку за 30 дней...")
            await export_operations_to_excel(callback, 30)
        except Exception as e:
            logger.error(f"Error exporting operations for 30 days: {e}", exc_info=True)
            await callback.message.answer("❌ Произошла ошибка при формировании выгрузки.")
            await callback.answer()
        keyboard = build_operations_history_keyboard()
        log_history_keyboard(callback, keyboard, source="operations_history_30")
        await callback.message.edit_reply_markup(reply_markup=keyboard)
        return
    else:
        days = 1  # Default to 1 day
    
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, callback.from_user.id)
        # Get operations - limit to avoid MESSAGE_TOO_LONG error
        # Telegram has a limit of 4096 characters per message
        if days == 1:
            # Show up to 30 operations for 1 day (to avoid message too long)
            operations = BillingService.get_user_operations(db, user.id, limit=30, days=days)
        elif days is None:
            # For "all" view, limit to 20
            operations = BillingService.get_user_operations(db, user.id, limit=20, days=days)
        else:
            # Should not happen (7 and 30 days are handled separately)
            operations = BillingService.get_user_operations(db, user.id, limit=20, days=days)
        total_count = BillingService.get_operations_count(db, user.id, days=days)
        
        if not operations:
            text = (
                "📊 **История операций**\n\n"
                "У вас пока нет операций.\n"
                "История появится после выполнения операций."
            )
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="↩️ Назад",
                        callback_data="payment_menu"
                    )
                ]
            ])
            await callback.message.edit_text(text, reply_markup=keyboard, parse_mode="Markdown")
            await callback.answer()
            return
        
        # Format operations history header
        period_text = ""
        if days == 1:
            period_text = " (за 1 день)"
        elif days == 7:
            period_text = " (за 7 дней)"
        elif days == 30:
            period_text = " (за 30 дней)"
        
        lines = [f"📊 **История операций{period_text}**\nВсего: {total_count}\n"]
        
        # Status emoji mapping
        status_emoji = {
            "charged": "✅",
            "pending": "⏳",
            "failed": "❌",
            "free": "🎁",
            "refunded": "↩️",
        }
        
        # Status text mapping
        status_text = {
            "charged": "Списано",
            "pending": "Ожидает",
            "failed": "Ошибка",
            "free": "Бесплатно",
            "refunded": "Возврат",
        }
        
        # Operation type emoji mapping
        type_emoji = {
            "generate": "🎨",
            "edit": "✏️",
            "merge": "✏️",
            "retouch": "✨",
            "upscale": "🔍",
            "prompt_generation": "✍️",
            "face_swap": "🔄",
            "add_text": "📝",
            "payment": "💰",  # Payment/deposit
        }
        
        # Limit operations to avoid MESSAGE_TOO_LONG error
        # Show up to 30 for 1 day, 20 for "all" view
        max_operations = 30 if days == 1 else 20
        operations_to_show = operations[:max_operations]
        
        # Build message and check length, reduce if needed
        # Telegram limit is 4096 characters, but we'll use 3500 to be safe
        MAX_MESSAGE_LENGTH = 3500
        
        for op in operations_to_show:
            op_type = op["type"]
            record_type = op.get("record_type", "operation")
            
            # Handle payment records
            if record_type == "payment" or op_type == "payment":
                op_name = "Пополнение баланса"
                type_icon = "💰"
                status = op["status"]
                # For payments, show as succeeded
                emoji = "✅"
                status_label = "Пополнено"
            else:
                op_name = get_operation_name(op_type)
                type_icon = type_emoji.get(op_type, "•")
                status = op["status"]
                emoji = status_emoji.get(status, "•")
                status_label = status_text.get(status, status)
            
            # Format date in Moscow timezone (compact format)
            created_at = op["created_at"]
            if isinstance(created_at, datetime):
                # Convert to Moscow timezone (UTC+3)
                moscow_tz = ZoneInfo("Europe/Moscow")
                # If datetime is naive (no timezone), assume it's UTC
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=ZoneInfo("UTC"))
                # Convert to Moscow time
                moscow_time = created_at.astimezone(moscow_tz)
                # Compact date format: DD.MM HH:MM
                date_str = moscow_time.strftime("%d.%m %H:%M")
            else:
                date_str = str(created_at)[:11]  # Just date part
            
            # Format price with discount info if available
            # Prices are stored in kopecks, convert to rubles for display
            price_rubles = op['price'] / 100.0
            original_price_kopecks = op.get("original_price")
            discount_percent = op.get("discount_percent")
            
            if op['price'] > 0:
                if original_price_kopecks and discount_percent and original_price_kopecks > op['price']:
                    # Show discount info
                    original_price_rubles = original_price_kopecks / 100.0
                    discount_amount_rubles = (original_price_kopecks - op['price']) / 100.0
                    price_str = (
                        f"~~{original_price_rubles:.2f} ₽~~ {price_rubles:.2f} ₽ "
                        f"🎟️ (скидка {discount_percent}%, -{discount_amount_rubles:.2f} ₽)"
                    )
                else:
                    price_str = f"{price_rubles:.2f} ₽"
            else:
                price_str = "Бесплатно"
            
            # Compact format: one line per operation
            # Handle payment records (always show)
            if record_type == "payment" or op_type == "payment":
                lines.append(f"{type_icon} {op_name} • {emoji} +{price_str} • {date_str}")
            # Only show charged, free, failed, or refunded operations in history
            # PENDING operations are not shown (they haven't been charged yet)
            elif status == "charged" or status == "free":
                lines.append(f"{type_icon} {op_name} • {emoji} {price_str} • {date_str}")
            elif status == "failed":
                lines.append(f"{type_icon} {op_name} • {emoji} {status_label} • {date_str}")
            elif status == "refunded":
                lines.append(f"{type_icon} {op_name} • {emoji} {price_str} • {status_label} • {date_str}")
            # PENDING operations are skipped - they haven't been charged yet
        
        # Show "... и еще" message if there are more operations than displayed
        displayed_count = len(operations_to_show)
        if total_count > displayed_count:
            remaining = total_count - displayed_count
            lines.append(f"\n... и еще {remaining} операций")
            lines.append("💡 Для полной выгрузки используйте кнопки «7 дней (Excel)» или «30 дней (Excel)»")
        
        text = "\n".join(lines)
        
        # Check message length and reduce if needed
        MAX_MESSAGE_LENGTH = 3500
        if len(text) > MAX_MESSAGE_LENGTH:
            # Reduce operations until message fits
            logger.warning(f"Message too long ({len(text)} chars), reducing operations")
            while len(text) > MAX_MESSAGE_LENGTH and len(operations_to_show) > 5:
                operations_to_show = operations_to_show[:-1]
                # Rebuild lines
                lines = [f"📊 **История операций{period_text}**\nВсего: {total_count}\n"]
                for op in operations_to_show:
                    op_type = op["type"]
                    record_type = op.get("record_type", "operation")
                    
                    if record_type == "payment" or op_type == "payment":
                        op_name = "Пополнение баланса"
                        type_icon = "💰"
                        emoji = "✅"
                        price_rubles = op['price'] / 100.0
                        price_str = f"{price_rubles:.2f} ₽"
                    else:
                        op_name = get_operation_name(op_type)
                        type_icon = type_emoji.get(op_type, "•")
                        status = op["status"]
                        emoji = status_emoji.get(status, "•")
                        price_rubles = op['price'] / 100.0
                        original_price_kopecks = op.get("original_price")
                        discount_percent = op.get("discount_percent")
                        
                        if op['price'] > 0:
                            if original_price_kopecks and discount_percent and original_price_kopecks > op['price']:
                                original_price_rubles = original_price_kopecks / 100.0
                                discount_amount_rubles = (original_price_kopecks - op['price']) / 100.0
                                price_str = f"~~{original_price_rubles:.2f}₽~~ {price_rubles:.2f}₽ 🎟️"
                            else:
                                price_str = f"{price_rubles:.2f}₽"
                        else:
                            price_str = "Бесплатно"
                    
                    created_at = op["created_at"]
                    if isinstance(created_at, datetime):
                        moscow_tz = ZoneInfo("Europe/Moscow")
                        if created_at.tzinfo is None:
                            created_at = created_at.replace(tzinfo=ZoneInfo("UTC"))
                        moscow_time = created_at.astimezone(moscow_tz)
                        date_str = moscow_time.strftime("%d.%m %H:%M")
                    else:
                        date_str = str(created_at)[:11]
                    
                    if record_type == "payment" or op_type == "payment":
                        lines.append(f"{type_icon} {op_name} • {emoji} +{price_str} • {date_str}")
                    elif status == "charged" or status == "free":
                        lines.append(f"{type_icon} {op_name} • {emoji} {price_str} • {date_str}")
                
                if total_count > len(operations_to_show):
                    remaining = total_count - len(operations_to_show)
                    lines.append(f"\n... и еще {remaining} операций")
                    if days == 1:
                        lines.append("💡 Для полной выгрузки используйте Excel")
                
                text = "\n".join(lines)
        
        keyboard = build_operations_history_keyboard()
        log_history_keyboard(callback, keyboard, source=f"operations_history_{days}")
        await callback.message.edit_text(text, reply_markup=keyboard, parse_mode="Markdown")
        await callback.answer()
    finally:
        db.close()


@router.callback_query(F.data == "operations_history_all")
async def legacy_operations_history_all(callback: CallbackQuery):
    """
    Легаси-обработчик для старых кнопок 'Все' (operations_history_all).
    Новые клавиатуры эту кнопку не создают, но старые сообщения продолжают слать этот callback.
    Просто показываем новое меню выбора периода.
    """
    keyboard = build_operations_history_keyboard()
    log_history_keyboard(callback, keyboard, source="operations_history_all_legacy")
    await callback.message.edit_text(
        "📊 Выберите период для просмотра истории операций:",
        reply_markup=keyboard,
        parse_mode="Markdown",
    )
    await callback.answer()


def check_balance_decorator(operation_type: str):
    """
    Decorator to check balance before executing paid operation.
    
    Usage:
        @check_balance_decorator("generate")
        async def handler(...):
            ...
    """
    def decorator(func):
        async def wrapper(*args, **kwargs):
            # Extract message or callback from args
            message_or_callback = None
            for arg in args:
                if isinstance(arg, (Message, CallbackQuery)):
                    message_or_callback = arg
                    break
            
            if not message_or_callback:
                logger.error("Could not find Message or CallbackQuery in handler args")
                return await func(*args, **kwargs)

            user_id = message_or_callback.from_user.id
            
            db = SessionLocal()
            try:
                user, _ = BillingService.get_or_create_user(db, user_id, message_or_callback.from_user)
                
                # Try to charge operation
                success, error_msg, operation_id = BillingService.charge_operation(
                    db,
                    user.id,
                    operation_type
                )

                if not success:
                    # Insufficient balance
                    user_info = BillingService.get_user_info(db, user_id)
                    balance = user_info["balance"] if user_info else 0
                    
                    text = (
                        f"❌ **Недостаточно средств**\n\n"
                        f"Операция стоит: 10 ₽\n"
                        f"Ваш баланс: {format_balance(balance)} ₽\n\n"
                        f"Пополните баланс для продолжения работы."
                    )
                    
                    keyboard = InlineKeyboardMarkup(inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text="💰 Пополнить баланс",
                                callback_data="payment_menu"
                            )
                        ]
                    ])

                    if isinstance(message_or_callback, Message):
                        await message_or_callback.answer(text, reply_markup=keyboard, parse_mode="Markdown")
                    else:
                        await message_or_callback.message.answer(text, reply_markup=keyboard, parse_mode="Markdown")
                        await message_or_callback.answer()
                    
                    return

                # Store operation_id in kwargs for use in handler
                kwargs['operation_id'] = operation_id
                return await func(*args, **kwargs)

            finally:
                db.close()

        return wrapper
    return decorator


@router.callback_query(F.data == "payment_discount_code")
async def callback_payment_discount_code(callback: CallbackQuery, state: FSMContext):
    """Request discount code input for payment."""
    await state.set_state(PaymentStates.WAIT_DISCOUNT_CODE)
    logger.info(f"Set state to WAIT_DISCOUNT_CODE for user {callback.from_user.id}")
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text="❌ Отмена",
                callback_data="payment_cancel"
            )
        ]
    ])

    await callback.message.edit_text(
        "🎟️ **Введите промокод**\n\n"
        "Промокод даст скидку на пополнение баланса.\n"
        "Введите промокод:",
        reply_markup=keyboard,
        parse_mode="Markdown"
    )
    await callback.answer()


async def process_discount_code(message: Message, state: FSMContext):
    """Process discount code input for payment."""
    current_state = await state.get_state()
    logger.info(f"process_discount_code called: text='{message.text}', user_id={message.from_user.id if message.from_user else 'unknown'}, state={current_state}")
    
    if not message.text:
        logger.warning("process_discount_code: message.text is None")
        return
    
    code = message.text.strip().upper()
    
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
        is_valid, discount, error_msg = DiscountService.validate_discount_code(
            db, code, user.id
        )

        if not is_valid:
            logger.warning(f"process_discount_code: invalid code '{code}': {error_msg}")
            await message.answer(f"❌ {error_msg}")
            return

        # Store discount code in state and show payment menu
        await state.update_data(discount_code=code, discount_id=discount.id)
        
        user_info = BillingService.get_user_info(db, message.from_user.id)
        if not user_info:
            user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
            user_info = BillingService.get_user_info(db, message.from_user.id)

        balance = user_info["balance"]
        free_left = user_info["free_operations_left"]
        free_total = user_info["free_operations_total"]

        text = (
            f"✅ **Промокод применен!**\n\n"
            f"🎟️ Промокод: {code}\n"
            f"💰 Скидка: {discount.discount_percent}%\n\n"
            f"💰 **Ваш баланс:** {balance} ₽\n\n"
            f"Выберите сумму для пополнения:"
        )

        keyboard = build_payment_keyboard()
        await message.answer(
            text,
            reply_markup=keyboard,
            parse_mode="Markdown"
        )
        # Сбрасываем состояние, но сохраняем данные промокода для применения при оплате
        await state.set_state(None)
        logger.info(f"process_discount_code: successfully applied code '{code}', state cleared")
    except Exception as e:
        logger.error(f"process_discount_code: error processing code '{code}': {e}", exc_info=True)
        await message.answer("❌ Произошла ошибка при обработке промокода.")
    finally:
        db.close()


@router.callback_query(F.data == "operation_discount_code")
async def callback_operation_discount_code(callback: CallbackQuery, state: FSMContext):
    """Request discount code input for operations."""
    await state.set_state(OperationDiscountStates.WAIT_OPERATION_DISCOUNT_CODE)
    logger.info(f"Set state to WAIT_OPERATION_DISCOUNT_CODE for user {callback.from_user.id}")
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text="❌ Отмена",
                callback_data="operation_discount_cancel"
            )
        ]
    ])

    await callback.message.edit_text(
        "🎟️ **Введите промокод для операций**\n\n"
        "Промокод даст скидку на все операции (генерация, редактирование, ретушь и т.д.).\n"
        "Введите промокод:",
        reply_markup=keyboard,
        parse_mode="Markdown"
    )
    await callback.answer()


async def process_operation_discount_code(message: Message, state: FSMContext):
    """Process discount code input for operations."""
    current_state = await state.get_state()
    logger.info(f"process_operation_discount_code called: text='{message.text}', user_id={message.from_user.id if message.from_user else 'unknown'}, state={current_state}")
    
    if not message.text:
        logger.warning("process_operation_discount_code: message.text is None")
        return
    
    code = message.text.strip().upper()
    
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, message.from_user.id, message.from_user)
        is_valid, discount, error_msg = DiscountService.validate_discount_code(
            db, code, user.id
        )

        if not is_valid:
            logger.warning(f"process_operation_discount_code: invalid code '{code}': {error_msg}")
            await message.answer(f"❌ {error_msg}")
            return

        # Handle FREE_ACCESS code separately (activates unlimited free operations)
        if code == "FREE_ACCESS":
            success, error_msg = DiscountService.activate_free_access(db, discount, user.id)
            if not success:
                await message.answer(f"❌ {error_msg}")
                return
            
            user_info = BillingService.get_user_info(db, message.from_user.id)
            balance = user_info["balance"] if user_info else 0
            
            text = (
                f"✅ **Промокод FREE_ACCESS активирован!**\n\n"
                f"🎟️ Промокод: {code}\n"
                f"✨ **Бесплатный доступ:** Активен\n\n"
                f"💡 Все операции теперь бесплатны!\n\n"
                f"💰 **Ваш баланс:** {format_balance(balance)} ₽"
            )
            
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="💰 Баланс",
                        callback_data="payment_menu"
                    )
                ]
            ])
        else:
            # Store discount code in database for persistence across restarts
            user.operation_discount_code_id = discount.id
            user.operation_discount_percent = discount.discount_percent
            db.commit()
            
            # Also store in state for immediate use
            await state.update_data(
                operation_discount_code=code,
                operation_discount_id=discount.id,
                operation_discount_percent=discount.discount_percent
            )
            
            user_info = BillingService.get_user_info(db, message.from_user.id)
            balance = user_info["balance"] if user_info else 0

            text = (
                f"✅ **Промокод применен к операциям!**\n\n"
                f"🎟️ Промокод: {code}\n"
                f"💰 Скидка: {discount.discount_percent}%\n\n"
                f"💡 Промокод будет применяться ко всем операциям до отмены.\n\n"
                f"💰 **Ваш баланс:** {format_balance(balance)} ₽"
            )
            
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="❌ Отменить промокод",
                        callback_data="operation_discount_remove"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="💰 Баланс",
                        callback_data="payment_menu"
                    )
                ]
            ])
        
        await message.answer(
            text,
            reply_markup=keyboard,
            parse_mode="Markdown"
        )
        await state.set_state(None)
        logger.info(f"process_operation_discount_code: successfully applied code '{code}', state cleared")
    except Exception as e:
        logger.error(f"process_operation_discount_code: error processing code '{code}': {e}", exc_info=True)
        await message.answer("❌ Произошла ошибка при обработке промокода.")
    finally:
        db.close()


@router.callback_query(F.data == "operation_discount_cancel")
async def callback_operation_discount_cancel(callback: CallbackQuery, state: FSMContext):
    """Cancel discount code input."""
    await state.set_state(None)
    await callback.message.edit_text(
        "❌ Ввод промокода отменен.",
        reply_markup=None
    )
    await callback.answer()


@router.callback_query(F.data == "operation_discount_remove")
async def callback_operation_discount_remove(callback: CallbackQuery, state: FSMContext):
    """Remove active discount code for operations."""
    db = SessionLocal()
    try:
        user, _ = BillingService.get_or_create_user(db, callback.from_user.id, callback.from_user)
        
        # Remove from database
        code = None
        if user.operation_discount_code_id:
            # Get discount code name for display
            from app.db.models import DiscountCode
            discount = db.query(DiscountCode).filter(DiscountCode.id == user.operation_discount_code_id).first()
            if discount:
                code = discount.code
            
            user.operation_discount_code_id = None
            user.operation_discount_percent = None
            db.commit()
        
        # Also remove from state
        await state.update_data(
            operation_discount_code=None,
            operation_discount_id=None,
            operation_discount_percent=None
        )
        
        if code:
            await callback.message.edit_text(
                f"✅ Промокод {code} отменен.\n\n"
                f"Теперь операции будут выполняться по обычным ценам.",
                reply_markup=None
            )
            logger.info(f"Operation discount code removed for user {callback.from_user.id}")
        else:
            await callback.message.edit_text(
                "❌ Активный промокод не найден.",
                reply_markup=None
            )
    except Exception as e:
        logger.error(f"Error removing operation discount code: {e}", exc_info=True)
        await callback.message.edit_text(
            "❌ Произошла ошибка при отмене промокода.",
            reply_markup=None
        )
    finally:
        db.close()
    await callback.answer()


async def handle_text_after_balance_menu(message: Message, state: FSMContext):
    """Handle text input after balance menu was shown (intercept before image handler)."""
    if not message.text:
        return
    
    # Check if balance menu was shown
    current_state = await state.get_state()
    logger.info("handle_text_after_balance_menu called: user_id={}, text='{}', state={}", 
               message.from_user.id if message.from_user else None,
               message.text,
               current_state)
    
    # Если пользователь явно пытается создать изображение (выбрал модель), не перехватываем
    # Проверяем, есть ли в состоянии выбранная модель
    data = await state.get_data()
    selected_model = data.get("selected_model") or data.get("model")
    if selected_model:
        # Пользователь уже выбрал модель, значит он создает изображение - не перехватываем
        logger.info("handle_text_after_balance_menu: user has selected model '{}', skipping interception", selected_model)
        return
    
    if current_state == PaymentStates.BALANCE_MENU_SHOWN:
        # User entered text after seeing balance menu
        # Check if it's a number (payment amount)
        try:
            amount = int(message.text.strip())
            if 10 <= amount <= 100000:
                # It's a valid payment amount
                await message.answer(
                    "💳 **Для пополнения баланса используйте меню**\n\n"
                    "Нажмите кнопку «💰 Пополнить баланс» в меню баланса.\n"
                    "Затем выберите сумму или нажмите «🔢 Другая сумма» для ввода произвольной суммы.",
                    parse_mode="Markdown"
                )
                # Don't clear state - keep it active to continue intercepting text
                return  # Handled, stop processing
        except (ValueError, AttributeError):
            # Not a number, but still show hint
            pass
        
        # Any text after balance menu - show hint
        # Don't clear state - keep it active to continue intercepting text
        await message.answer(
            "💡 **Используйте кнопки меню для работы с балансом**\n\n"
            "Для пополнения баланса нажмите «💰 Пополнить баланс».\n"
            "Для просмотра истории операций нажмите «📊 История операций».\n\n"
            "Если вы хотите создать изображение, используйте кнопку «🎨 Создать».",
            parse_mode="Markdown"
        )
        return  # Handled, stop processing


def register_billing_handlers(dp):
    """Register billing handlers to dispatcher."""
    # Регистрируем обработчик кнопки баланса через dp.message.register с высоким приоритетом
    # Это гарантирует, что он будет проверяться раньше общего обработчика текста в image.py
    logger.info("Registering balance button handler: BALANCE_BUTTON='{}'", BALANCE_BUTTON)
    dp.message.register(handle_balance_button, F.text == BALANCE_BUTTON)
    logger.info("Balance button handler registered successfully")
    
    # Регистрируем обработчик текста после показа меню баланса
    # Этот обработчик должен быть зарегистрирован ПОСЛЕ image handlers
    # чтобы он проверялся ПЕРВЫМ (в aiogram обработчики проверяются в обратном порядке)
    # Регистрируем его здесь, но он будет вызван из __init__.py после image handlers
    logger.info("Text after balance menu handler registration skipped here, will be registered after image handlers")
    
    # Регистрируем обработчик промокода с высоким приоритетом
    # В aiogram обработчики проверяются в обратном порядке регистрации (последний = первый)
    # Поэтому регистрируем ПОСЛЕ image handlers, чтобы он проверялся ПЕРВЫМ
    logger.info("Registering discount code handler with high priority")
    dp.message.register(
        process_discount_code,
        StateFilter(PaymentStates.WAIT_DISCOUNT_CODE),
        F.text
    )
    logger.info("Discount code handler registered successfully")
    
    # Регистрируем обработчик кастомной суммы с высоким приоритетом
    logger.info("Registering custom amount handler with high priority")
    dp.message.register(
        process_custom_amount,
        StateFilter(PaymentStates.WAIT_CUSTOM_AMOUNT),
        F.text
    )
    logger.info("Custom amount handler registered successfully")
    
    # Регистрируем обработчик email с высоким приоритетом
    logger.info("Registering email handler with high priority")
    dp.message.register(
        process_email,
        StateFilter(PaymentStates.WAIT_EMAIL),
        F.text
    )
    logger.info("Email handler registered successfully")
    
    # Регистрируем обработчик промокода для операций с высоким приоритетом
    logger.info("Registering operation discount code handler with high priority")
    dp.message.register(
        process_operation_discount_code,
        StateFilter(OperationDiscountStates.WAIT_OPERATION_DISCOUNT_CODE),
        F.text
    )
    logger.info("Operation discount code handler registered successfully")
    
    # Регистрируем остальные handlers через router
    dp.include_router(router)


@router.message(Command("export_stats"))
async def handle_export_stats(message: Message):
    """Export statistics to Excel file."""
    try:
        import tempfile
        import os
        from scripts.export_statistics_to_excel import export_statistics_to_excel
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            excel_file = tmp.name
        
        try:
            # Export statistics using the centralized script
            export_statistics_to_excel(excel_file)
            
            # Send file to user
            from aiogram.types import FSInputFile
            await message.answer("📊 Готовлю выгрузку статистики...")
            
            file = FSInputFile(excel_file, filename="statistics_export.xlsx")
            await message.answer_document(
                document=file,
                caption=f"📊 Статистика на {get_moscow_time().strftime('%d.%m.%Y %H:%M')}"
            )
        finally:
            # Clean up temporary file
            if os.path.exists(excel_file):
                os.unlink(excel_file)
        
        return
    except Exception as e:
        logger.error(f"Error in handle_export_stats: {e}", exc_info=True)
        await message.answer("❌ Произошла ошибка при экспорте статистики.")
    
    # OLD CODE BELOW - REMOVED, USING CENTRALIZED SCRIPT INSTEAD
    """
    db = SessionLocal()
    try:
        from pathlib import Path
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from app.db.models import User, UserStatistics, Operation, Balance
        from app.services.pricing import get_operation_name
        from sqlalchemy import func, desc
        import json
        from datetime import datetime
        
        # Create temporary file
        temp_file = Path("/tmp/statistics_export.xlsx")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Users sheet
        ws_users = wb.create_sheet("Пользователи")
        headers = ["ID", "Telegram ID", "Username", "Имя", "Фамилия", "Язык", "Premium", 
                  "Регистрация", "Последняя активность", "Баланс", "Всего операций", 
                  "Всего потрачено", "Первая операция", "Последняя операция"]
        ws_users.append(headers)
        
        # Style headers
        for cell in ws_users[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        users = db.query(User).order_by(desc(User.created_at)).all()
        
        for user in users:
            stats = db.query(UserStatistics).filter(UserStatistics.user_id == user.id).first()
            balance = db.query(Balance).filter(Balance.user_id == user.id).first()
            
            ws_users.append([
                user.id,
                user.telegram_id,
                f"@{user.username}" if user.username else "",
                user.first_name or "",
                user.last_name or "",
                user.language_code or "",
                "Да" if user.is_premium else "Нет",
                user.created_at.strftime("%d.%m.%Y %H:%M") if user.created_at else "",
                user.last_activity_at.strftime("%d.%m.%Y %H:%M") if user.last_activity_at else "",
                balance.balance if balance else 0,
                stats.total_operations if stats else 0,
                stats.total_spent if stats else 0,
                stats.first_operation_at.strftime("%d.%m.%Y %H:%M") if stats and stats.first_operation_at else "",
                stats.last_operation_at.strftime("%d.%m.%Y %H:%M") if stats and stats.last_operation_at else "",
            ])
        
        # 2. Operations by type sheet
        ws_ops_type = wb.create_sheet("Операции по типам")
        ws_ops_type.append(["Тип операции", "Количество", "Выручка (₽)"])
        for cell in ws_ops_type[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        ops_by_type_query = db.query(
            Operation.type,
            func.count(Operation.id).label('count'),
            func.sum(Operation.price).label('total_revenue')
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).group_by(Operation.type).order_by(desc('count')).all()
        
        for op_type, count, revenue in ops_by_type_query:
            ws_ops_type.append([get_operation_name(op_type), count, revenue or 0])
        
        # 3. Models used sheet
        ws_models = wb.create_sheet("Использованные модели")
        ws_models.append(["Модель", "Количество использований", "Выручка (₽)"])
        for cell in ws_models[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        models_query = db.query(
            Operation.model,
            func.count(Operation.id).label('count'),
            func.sum(Operation.price).label('total_revenue')
        ).filter(
            Operation.status.in_(["charged", "free"]),
            Operation.model.isnot(None)
        ).group_by(Operation.model).order_by(desc('count')).all()
        
        for model, count, revenue in models_query:
            ws_models.append([model, count, revenue or 0])
        
        # 4. All operations sheet
        ws_operations = wb.create_sheet("Все операции")
        ws_operations.append(["ID операции", "Telegram ID", "Тип", "Модель", "Цена", 
                             "Статус", "Дата", "Промпт", "Количество изображений"])
        for cell in ws_operations[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        operations = db.query(Operation).filter(
            Operation.status.in_(["charged", "free"])
        ).order_by(desc(Operation.created_at)).all()
        
        for op in operations:
            user = db.query(User).filter(User.id == op.user_id).first()
            prompt = (op.prompt[:200] + "...") if op.prompt and len(op.prompt) > 200 else (op.prompt or "")
            ws_operations.append([
                op.id,
                user.telegram_id if user else "",
                get_operation_name(op.type),
                op.model or "",
                op.price,
                op.status,
                op.created_at.strftime("%d.%m.%Y %H:%M") if op.created_at else "",
                prompt,
                op.image_count or ""
            ])
        
        # 5. Summary sheet
        ws_summary = wb.create_sheet("Сводка")
        ws_summary.append(["Параметр", "Значение"])
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        total_users = db.query(func.count(User.id)).scalar()
        total_operations = db.query(func.count(Operation.id)).filter(
            Operation.status.in_(["charged", "free"])
        ).scalar()
        total_revenue = db.query(func.sum(Operation.price)).filter(
            Operation.status == "charged"
        ).scalar() or 0
        total_balance = db.query(func.sum(Balance.balance)).scalar() or 0
        
        ws_summary.append(["Всего пользователей", total_users])
        ws_summary.append(["Всего операций", total_operations])
        ws_summary.append(["Всего заработано (₽)", total_revenue])
        ws_summary.append(["Общий баланс пользователей (₽)", total_balance])
        ws_summary.append(["Дата выгрузки", get_moscow_time().strftime("%d.%m.%Y %H:%M")])
        
        # 6. User operations statistics sheet
        from collections import defaultdict
        from datetime import timedelta
        
        ws_user_ops = wb.create_sheet("Статистика по пользователям")
        ws_user_ops.append(["Telegram ID", "Username", "Имя", "Тип операции", "Количество", "Выручка (₽)"])
        for cell in ws_user_ops[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get operations grouped by user and type
        user_ops_query = db.query(
            User.telegram_id,
            User.username,
            User.first_name,
            Operation.type,
            func.count(Operation.id).label('count'),
            func.sum(Operation.price).label('revenue')
        ).join(
            Operation, User.id == Operation.user_id
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).group_by(
            User.telegram_id, User.username, User.first_name, Operation.type
        ).order_by(User.telegram_id, desc('count')).all()
        
        for tg_id, username, first_name, op_type, count, revenue in user_ops_query:
            ws_user_ops.append([
                tg_id,
                f"@{username}" if username else "",
                first_name or "",
                get_operation_name(op_type),
                count,
                revenue or 0
            ])
        
        # 7. Daily statistics sheet
        ws_daily = wb.create_sheet("Статистика по дням")
        ws_daily.append(["Дата", "Количество операций", "Выручка (₽)", "Уникальных пользователей"])
        for cell in ws_daily[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        daily_stats = db.query(
            func.date(Operation.created_at).label('date'),
            func.count(Operation.id).label('count'),
            func.sum(Operation.price).label('revenue'),
            func.count(func.distinct(Operation.user_id)).label('unique_users')
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).group_by(
            func.date(Operation.created_at)
        ).order_by(desc('date')).all()
        
        for date, count, revenue, unique_users in daily_stats:
            ws_daily.append([
                date.strftime("%d.%m.%Y") if isinstance(date, datetime) else str(date),
                count,
                revenue or 0,
                unique_users
            ])
        
        # 8. Weekly statistics sheet
        ws_weekly = wb.create_sheet("Статистика по неделям")
        ws_weekly.append(["Неделя", "Количество операций", "Выручка (₽)", "Уникальных пользователей"])
        for cell in ws_weekly[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get week start dates (Monday)
        weekly_data = defaultdict(lambda: {'count': 0, 'revenue': 0, 'users': set()})
        
        ops_for_weekly = db.query(
            Operation.created_at,
            Operation.price,
            Operation.user_id
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).all()
        
        for op in ops_for_weekly:
            if op.created_at:
                # Get Monday of the week
                week_start = op.created_at - timedelta(days=op.created_at.weekday())
                week_key = week_start.strftime("%d.%m.%Y")
                
                weekly_data[week_key]['count'] += 1
                weekly_data[week_key]['revenue'] += op.price or 0
                weekly_data[week_key]['users'].add(op.user_id)
        
        # Sort by date descending
        sorted_weeks = sorted(weekly_data.items(), key=lambda x: datetime.strptime(x[0], "%d.%m.%Y"), reverse=True)
        
        for week_key, data in sorted_weeks:
            # Calculate week end (Sunday)
            week_start = datetime.strptime(week_key, "%d.%m.%Y")
            week_end = week_start + timedelta(days=6)
            ws_weekly.append([
                f"{week_key} - {week_end.strftime('%d.%m.%Y')}",
                data['count'],
                data['revenue'],
                len(data['users'])
            ])
        
        # 9. Monthly statistics sheet
        ws_monthly = wb.create_sheet("Статистика по месяцам")
        ws_monthly.append(["Месяц", "Количество операций", "Выручка (₽)", "Уникальных пользователей"])
        for cell in ws_monthly[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        from sqlalchemy import extract
        monthly_stats = db.query(
            extract('year', Operation.created_at).label('year'),
            extract('month', Operation.created_at).label('month'),
            func.count(Operation.id).label('count'),
            func.sum(Operation.price).label('revenue'),
            func.count(func.distinct(Operation.user_id)).label('unique_users')
        ).filter(
            Operation.status.in_(["charged", "free"])
        ).group_by(
            extract('year', Operation.created_at),
            extract('month', Operation.created_at)
        ).order_by(
            desc('year'), desc('month')
        ).all()
        
        # Russian month names
        month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
        }
        
        for year, month, count, revenue, unique_users in monthly_stats:
            month_name = f"{month_names.get(int(month), str(month))} {int(year)}"
            ws_monthly.append([
                month_name,
                count,
                revenue or 0,
                unique_users
            ])
        
        wb.save(temp_file)
        
        # Send file to user
        from aiogram.types import FSInputFile
        await message.answer("📊 Готовлю выгрузку статистики...")
        
        file = FSInputFile(temp_file)
        await message.answer_document(
            document=file,
            caption=f"📊 Статистика на {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
        
        # Clean up
        if temp_file.exists():
            temp_file.unlink()
        
    except Exception as e:
        logger.error(f"Error in handle_export_stats: {e}", exc_info=True)
        await message.answer("❌ Произошла ошибка при экспорте статистики.")
    finally:
        db.close()
    """

