#!/usr/bin/env python3
"""Analyze data and test export logic."""
import sys
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.db.base import SessionLocal
from app.db.models import Operation, OperationStatus, User
from sqlalchemy import func, desc
from collections import defaultdict

KOPECKS_MIGRATION_DATETIME = datetime(2025, 11, 25, 9, 30, 0, tzinfo=timezone.utc)

db = SessionLocal()
try:
    print("=" * 80)
    print("АНАЛИЗ ДАННЫХ В БД")
    print("=" * 80)
    
    # 1. Check all operations
    all_ops = db.query(Operation).filter(
        Operation.status.in_([OperationStatus.CHARGED, OperationStatus.FREE])
    ).order_by(Operation.created_at).all()
    
    print(f"\nВсего операций: {len(all_ops)}")
    
    # Count operations before and after migration
    ops_before = []
    ops_after = []
    
    for op in all_ops:
        if op.created_at:
            if op.created_at.tzinfo is None:
                created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
            else:
                created_at_utc = op.created_at.astimezone(timezone.utc)
            
            if created_at_utc < KOPECKS_MIGRATION_DATETIME:
                ops_before.append(op)
            else:
                ops_after.append(op)
        elif op.price > 100:
            ops_after.append(op)
        else:
            ops_before.append(op)
    
    print(f"Операций до миграции (игнорируем): {len(ops_before)}")
    print(f"Операций после миграции (учитываем): {len(ops_after)}")
    
    # 2. Calculate revenue from operations after migration
    total_revenue = 0.0
    ops_by_type = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    
    print("\n" + "=" * 80)
    print("РАСЧЕТ ВЫРУЧКИ (только операции после миграции)")
    print("=" * 80)
    
    for op in ops_after:
        if op.status == OperationStatus.CHARGED:
            if op.created_at:
                if op.created_at.tzinfo is None:
                    created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                else:
                    created_at_utc = op.created_at.astimezone(timezone.utc)
                
                if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                    price_rubles = float(op.price) / 100.0
                    total_revenue += price_rubles
                    ops_by_type[op.type]['revenue'] += price_rubles
                    ops_by_type[op.type]['count'] += 1
            elif op.price > 100:
                price_rubles = float(op.price) / 100.0
                total_revenue += price_rubles
                ops_by_type[op.type]['revenue'] += price_rubles
                ops_by_type[op.type]['count'] += 1
    
    print(f"\nОбщая выручка: {total_revenue:.2f} ₽")
    print("\nВыручка по типам операций:")
    for op_type, data in sorted(ops_by_type.items(), key=lambda x: -x[1]['revenue']):
        print(f"  {op_type}: {data['count']} операций, {data['revenue']:.2f} ₽")
    
    # 3. Check user statistics
    print("\n" + "=" * 80)
    print("СТАТИСТИКА ПО ПОЛЬЗОВАТЕЛЯМ (только операции после миграции)")
    print("=" * 80)
    
    users = db.query(User).all()
    for user in users:
        user_ops_after = [op for op in ops_after if op.user_id == user.id]
        user_revenue = 0.0
        user_count = 0
        
        for op in user_ops_after:
            if op.status == OperationStatus.CHARGED:
                if op.created_at:
                    if op.created_at.tzinfo is None:
                        created_at_utc = op.created_at.replace(tzinfo=timezone.utc)
                    else:
                        created_at_utc = op.created_at.astimezone(timezone.utc)
                    
                    if created_at_utc >= KOPECKS_MIGRATION_DATETIME:
                        price_rubles = float(op.price) / 100.0
                        user_revenue += price_rubles
                        user_count += 1
                elif op.price > 100:
                    price_rubles = float(op.price) / 100.0
                    user_revenue += price_rubles
                    user_count += 1
        
        if user_count > 0:
            print(f"\nПользователь {user.telegram_id} ({user.first_name or 'N/A'}):")
            print(f"  Операций: {user_count}")
            print(f"  Выручка: {user_revenue:.2f} ₽")
    
    # 4. Test export function
    print("\n" + "=" * 80)
    print("ТЕСТИРОВАНИЕ ЭКСПОРТА")
    print("=" * 80)
    
    from scripts.export_statistics_to_excel import export_statistics_to_excel
    import tempfile
    import os
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        test_file = tmp.name
    
    try:
        export_statistics_to_excel(test_file)
        print(f"\n✅ Экспорт выполнен успешно: {test_file}")
        print(f"Размер файла: {os.path.getsize(test_file)} байт")
        
        # Try to read and verify
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        
        print("\nЛисты в файле:")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  {sheet_name}: {ws.max_row} строк")
            
            # Check summary sheet
            if sheet_name == "Сводка":
                print("\n  Данные из листа 'Сводка':")
                for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                    if row[0] and row[1]:
                        print(f"    {row[0]}: {row[1]}")
        
    except Exception as e:
        print(f"\n❌ Ошибка при экспорте: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"\nВременный файл удален: {test_file}")
    
finally:
    db.close()





