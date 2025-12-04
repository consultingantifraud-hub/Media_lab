#!/usr/bin/env python3
"""Test profit calculation in export."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.export_statistics_to_excel import export_statistics_to_excel, get_model_cost_rub
from openpyxl import load_workbook
import tempfile
import os

# Test model cost calculation
print("Тест расчета стоимости моделей:")
print(f"  fal-ai/nano-banana-pro: {get_model_cost_rub('fal-ai/nano-banana-pro'):.2f} ₽")
print(f"  fal-ai/nano-banana: {get_model_cost_rub('fal-ai/nano-banana'):.2f} ₽")
print(f"  fal-ai/bytedance/seedream/v4/text-to-image: {get_model_cost_rub('fal-ai/bytedance/seedream/v4/text-to-image'):.2f} ₽")
print(f"  None: {get_model_cost_rub(None):.2f} ₽")
print()

# Create test export
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    test_file = tmp.name

try:
    print("Создание экспорта...")
    export_statistics_to_excel(test_file)
    
    print(f"\nЧтение файла: {test_file}")
    wb = load_workbook(test_file)
    
    # Check each sheet
    sheets_to_check = [
        "Использованные модели",
        "Статистика по пользователям",
        "Статистика по дням",
        "Статистика по неделям",
        "Статистика по месяцам",
        "Сводка"
    ]
    
    for sheet_name in sheets_to_check:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📊 {sheet_name}:")
            headers = [cell.value for cell in ws[1]]
            print(f"   Заголовки ({len(headers)}): {headers}")
            
            # Show first few data rows
            print("   Первые строки данных:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True), 2):
                if row and row[0]:
                    print(f"   {row_idx}: {row}")
        else:
            print(f"\n❌ Лист '{sheet_name}' не найден!")
    
finally:
    if os.path.exists(test_file):
        os.remove(test_file)





