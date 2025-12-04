#!/usr/bin/env python3
"""Check user operations sheet with models."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.export_statistics_to_excel import export_statistics_to_excel
from openpyxl import load_workbook
import tempfile
import os

# Create test export
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    test_file = tmp.name

try:
    print("Creating export...")
    export_statistics_to_excel(test_file)
    
    print(f"\nReading file: {test_file}")
    wb = load_workbook(test_file)
    
    # Check "Статистика по пользователям"
    if "Статистика по пользователям" in wb.sheetnames:
        ws = wb["Статистика по пользователям"]
        print("\n📊 Статистика по пользователям:")
        print(f"   Строк: {ws.max_row}, Колонок: {ws.max_column}")
        
        # Show headers
        headers = [cell.value for cell in ws[1]]
        print(f"\n   Заголовки: {headers}")
        
        # Show first 10 rows
        print("\n   Первые строки данных:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(11, ws.max_row), values_only=True), 2):
            if row and row[0]:
                print(f"   {row_idx}: {row}")
    else:
        print("\n❌ Лист 'Статистика по пользователям' не найден!")
    
finally:
    if os.path.exists(test_file):
        os.remove(test_file)





