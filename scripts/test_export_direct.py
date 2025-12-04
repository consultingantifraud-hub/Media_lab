#!/usr/bin/env python3
"""Test export directly."""
import sys
from pathlib import Path
import tempfile
import os

sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.export_statistics_to_excel import export_statistics_to_excel
from openpyxl import load_workbook

# Create test export
with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    test_file = tmp.name

try:
    print("Creating export...")
    export_statistics_to_excel(test_file)
    
    print(f"\nReading file: {test_file}")
    wb = load_workbook(test_file)
    
    # Check "Операции по типам"
    if "Операции по типам" in wb.sheetnames:
        ws = wb["Операции по типам"]
        print("\n📊 Операции по типам:")
        total_count = 0
        total_revenue = 0.0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row and row[0]:
                print(f"  {row[0]}: {row[1]} операций, {row[2]} ₽")
                if isinstance(row[1], (int, float)):
                    total_count += int(row[1])
                if isinstance(row[2], (int, float)):
                    total_revenue += float(row[2])
        print(f"\n  ИТОГО: {total_count} операций, {total_revenue:.2f} ₽")
    
    # Check "Использованные модели"
    if "Использованные модели" in wb.sheetnames:
        ws = wb["Использованные модели"]
        print("\n📊 Использованные модели:")
        for row in ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True):
            if row and row[0]:
                print(f"  {row[0]}: {row[1]} использований, {row[2]} ₽")
    
    # Check "Сводка"
    if "Сводка" in wb.sheetnames:
        ws = wb["Сводка"]
        print("\n📊 Сводка:")
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and row[0] and row[1]:
                print(f"  {row[0]}: {row[1]}")
    
finally:
    if os.path.exists(test_file):
        os.remove(test_file)





